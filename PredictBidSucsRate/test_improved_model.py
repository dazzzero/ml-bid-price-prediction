# -*- coding: utf-8 -*-
"""
개선된 데이터로 학습한 모델의 실제 성능 테스트 스크립트
bid.ml.train.py와 동일한 방식으로 결과를 엑셀로 생성

@author: user
"""

import os
import joblib
import numpy as np
import pandas as pd
from time import time
from sklearn.metrics import mean_squared_error, r2_score, mean_absolute_error
from sklearn.preprocessing import StandardScaler
from sklearn.model_selection import train_test_split
import warnings
warnings.filterwarnings('ignore')

# 엑셀 파일 처리를 위한 라이브러리
try:
    import openpyxl
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("⚠️  openpyxl이 설치되지 않았습니다. 엑셀 통계 기능을 사용하려면 'pip install openpyxl'을 실행하세요.")

# 한국어 자연어 처리를 위한 라이브러리
from kiwipiepy import Kiwi
from sklearn.feature_extraction.text import TfidfVectorizer
from scipy.sparse import csr_matrix

# 고급 특성 엔지니어링
from advanced_feature_engineering import AdvancedFeatureEngineering

class KiwiTokenizer():
    """한국어 텍스트를 처리하는 클래스"""
    
    def __init__(self, saved_filenm):
        self.rnd_num = np.random.randint(100, 999)
        self.cur_dir = os.getcwd()
        self.data_dir = self.cur_dir+'\\data\\'
        self.save_dir = self.cur_dir+'\\res\\'
        self.save_filename = saved_filenm
        self.kiwi = None
        self.kiwi = self.CreateKiwi(saved_filenm)
    
    def save(self, filename):
        joblib.dump(self.kiwi._user_values, self.save_dir+filename)
        return self
    
    def load(self, filename):
        o = joblib.load(filename)
        return o
        
    def loadDictonary(self, filename):
        self.data = pd.read_csv(self.data_dir+filename)
        for i, row in enumerate(self.data.iloc):
            self.kiwi.add_user_word(row["단어"], 'NNP')
    
    def CreateKiwi(self, saved_filenm):
        o = None
        if(self.kiwi is None):
            if(saved_filenm is None):
                o = Kiwi(num_workers=8)
            else:
                o = Kiwi(num_workers=8)
                o._user_values = self.load(saved_filenm)
        return o
            
    def cleared_line(self, line):
        if pd.isna(line) or line is None:
            line = ''
        else:
            line = str(line).lower().replace('(', ' ').replace(')', ' ').replace('n/a', '')
        
        nm_words = []
        tokens = self.kiwi.tokenize(line)
        for token in tokens:
            if token.tag in ['MM', 'NNG', 'NNB', 'NNP', 'SL', 'XPN', 'MAG', 'SN', 'SO', 'W_SERIAL']:
                nm_words.append(token.form)
                
        return ' '.join(nm_words)
    
    def cleared_lines_from(self, orglines):
        lines = []
        for line in orglines:
            lines.append(self.cleared_line(line))
        return lines      
    
    def nn_only(self, orglines):
        lines = []
        for key in orglines:
            if pd.isna(key) or key is None:
                key = ''
            else:
                key = str(key).lower().replace('(', ' ').replace(')', ' ').replace('n/a', '')
            
            nm_words = []
            tokens = self.kiwi.tokenize(key)
            for token in tokens:
                if token.tag in ['MM', 'NNG', 'NNB', 'NNP', 'SL', 'XPN', 'MAG', 'SN', 'SO', 'W_SERIAL']:
                    nm_words.append(token.form)
            lines.append(' '.join(nm_words))
            
        return lines

class KiwiVectorizer():
    def __init__(self):
        self.rnd_num = np.random.randint(100, 999)
        self.cur_dir = os.getcwd()
        self.save_dir = self.cur_dir+'\\res\\'
        
        self.vect = TfidfVectorizer(
            ngram_range=(1, 1), 
            token_pattern='(?u)\\b\\w+\\b',
            max_features=5000,
            min_df=2,
            max_df=0.95,
            sublinear_tf=True
        )
    
    def load(self, filename):
        voca = joblib.load(filename)
        self.vect.vocabulary_ = voca["vocabulary"] 
        self.vect.idf_ = voca["idf"]
    
    def save(self, filename):
        joblib.dump({'vocabulary':self.vect.vocabulary_, 'idf':self.vect.idf_}, self.save_dir + filename)
        print("단어사전 파일("+filename+")로 저장합니다. ")
    
    def fit(self, lines):
        return self.vect.fit(lines)
    
    def transform(self, lines):
        return self.vect.transform(lines)
    
    def scores(self, lines):
        return self.toValues(self.transform(lines))

    def toValues(self, csrmat:csr_matrix):
        lst = []
        safty_sz = len(csrmat.indptr)-1
        for i, n in enumerate(csrmat.indptr):
            sumval = 0
            if i<safty_sz:
                indiceVal = csrmat.indices[csrmat.indptr[i]:csrmat.indptr[i+1]]
                dataVal = csrmat.data[csrmat.indptr[i]:csrmat.indptr[i+1]]
                for j, va1 in enumerate(indiceVal):
                    sumval += va1 * dataVal[j]
                lst.append(sumval)
        return lst

class ImprovedModelTester():
    """개선된 모델의 성능을 테스트하는 클래스"""
    
    def __init__(self):
        self.cur_dir = os.getcwd()
        self.data_dir = self.cur_dir+'\\data\\'
        self.save_dir = self.cur_dir+'\\res\\'
        
        # 결과 컬럼 정의 (bid.ml.train.py와 동일)
        self.result_columns = ['입찰번호', '입찰차수', '기초금액', 
                               '낙찰하한률', '참여업체수', '간접비', '순공사원가', 
                               '면허제한코드', '공고기관코드',
                               '공고기관명', '공고기관점수',
                               '공사지역', '공사지역점수',                               
                               '키워드', '키워드점수',
                               '업체투찰률', '예가투찰률', '투찰률오차', 
                               '예정금액', '낙찰하한가', '낙찰금액', 
                               '업체투찰률예측', '예가투찰률예측', '참여업체수예측', '예정금액예측',
                               '낙찰금액(업체투찰률) 예측', 'A값여부', '결과1', 
                               '예정금액(예가투찰률) 예측', '예정금액*낙찰하한율', '결과2']
        
        # 결과 컬럼 타입 정의
        self.result_columns_type = {
                            '입찰번호':'str', '입찰차수':'int64','기초금액':'float64',
                            '낙찰하한률':'float64', '참여업체수':'float64', '간접비':'int64', '순공사원가':'int64',  
                            '면허제한코드':'float64', '공고기관코드':'float64',
                            '공고기관명':'str', '공고기관점수':'float64',
                            '공사지역':'str', '공사지역점수':'float64',                            
                            '키워드':'str', '키워드점수':'float64',
                            '업체투찰률':'float64', '예가투찰률':'float64', '투찰률오차':'float64', 
                            '예정금액':'int64', '낙찰하한가':'int64', '낙찰금액':'int64', 
                            '업체투찰률예측':'float64', '예가투찰률예측':'float64', '참여업체수예측':'float64', '예정금액예측':'int64',
                            '낙찰금액(업체투찰률) 예측':'float64', 'A값여부':'str', '결과1':'str',
                            '예정금액(예가투찰률) 예측':'float64', '예정금액*낙찰하한율':'float64', '결과2':'str'
                            }
        
        # 텍스트 처리 도구들 초기화
        self.tokenizer = KiwiTokenizer(self.save_dir + "mlpregr.tokenizer.v0.1.1.npz")
        self.vectorizer = KiwiVectorizer()
        self.vectorizer.load(self.save_dir + "mlpregr.vectorizer.v0.1.1.npz")
        
        # 스케일러 로드
        self.scaler = joblib.load(self.save_dir + 'x_fited_scaler.v2.npz')
        
        # 모델들 로드
        self.model1 = joblib.load(self.save_dir + 'mlpregr.model1.v0.1.1.npz')
        self.model2 = joblib.load(self.save_dir + 'mlpregr.model2.v0.1.1.npz')
        self.model3 = joblib.load(self.save_dir + 'mlpregr.model3.v0.1.1.npz')
        
        print("✅ 모델 및 전처리 도구 로드 완료")

    def load_test_data(self, filename):
        """테스트 데이터 로드 및 전처리"""
        print(f"📁 테스트 데이터 로드: {filename}")
        
        # 데이터 로드
        data = pd.read_csv(self.data_dir + filename, encoding='utf-8')
        print(f"✅ 데이터 로드 완료: {data.shape[0]}행, {data.shape[1]}열")
        
        # 필요한 컬럼들만 선택
        cvs_columns = ['기초금액', '낙찰하한률', '참여업체수', 
                      '낙찰금액', '업체투찰률', '예가투찰률', '투찰률오차', 
                      '간접비', '순공사원가', '입찰번호', '입찰차수', 
                      '예정금액', '낙찰하한가', 
                      '면허제한코드','공고기관코드','주공종명', 
                      '공고기관명', '공고기관점수',
                      '공사지역', '공사지역점수',
                      '키워드', '키워드점수']
        
        # 존재하는 컬럼만 선택
        available_columns = [col for col in cvs_columns if col in data.columns]
        dataset_x = pd.DataFrame(data, columns=available_columns)
        
        # 텍스트 컬럼 처리
        print("🔧 텍스트 데이터 처리 중...")
        if '키워드' in dataset_x.columns:
            dataset_x['키워드'] = dataset_x['키워드'].fillna('').astype(str)
        if '공고기관명' in dataset_x.columns:
            dataset_x['공고기관명'] = dataset_x['공고기관명'].fillna('').astype(str)
        if '공사지역' in dataset_x.columns:
            dataset_x['공사지역'] = dataset_x['공사지역'].fillna('').astype(str)
        
        # 문자열을 숫자로 변환
        if '면허제한코드' in dataset_x.columns:
            dataset_x['면허제한코드'] = dataset_x['면허제한코드'].astype(str).apply(
                lambda x: hash(x) % 1000000 if pd.notna(x) and x != 'nan' else 0)
        if '공고기관코드' in dataset_x.columns:
            dataset_x['공고기관코드'] = dataset_x['공고기관코드'].astype(str).apply(
                lambda x: hash(x) % 1000000 if pd.notna(x) and x != 'nan' else 0)
        
        # 텍스트 데이터 처리
        lines = self.tokenizer.nn_only(dataset_x["키워드"].tolist())
        lines2 = self.tokenizer.nn_only(dataset_x["공고기관명"].tolist())
        lines3 = self.tokenizer.nn_only(dataset_x["공사지역"].tolist())
        
        # TF-IDF 점수 계산
        pts = self.vectorizer.scores(lines)
        pts2 = self.vectorizer.scores(lines2)
        pts3 = self.vectorizer.scores(lines3)
        
        # 점수를 데이터에 추가
        dataset_x["키워드점수"] = pts
        dataset_x["공고기관점수"] = pts2
        dataset_x["공사지역점수"] = pts3
        
        # 고급 특성 엔지니어링 적용
        print("🔧 고급 특성 엔지니어링 적용 중...")
        feature_eng = AdvancedFeatureEngineering()
        
        dataset_x = feature_eng.create_interaction_features(dataset_x)
        dataset_x = feature_eng.create_ratio_features(dataset_x)
        dataset_x = feature_eng.create_categorical_features(dataset_x)
        dataset_x = feature_eng.create_statistical_features(dataset_x)
        
        print(f"✅ 특성 엔지니어링 완료: {dataset_x.shape[1]}개 특성")
        
        # 타겟 변수 준비
        dataset_y = pd.DataFrame(data, columns=['업체투찰률', '예가투찰률', '참여업체수'])
        
        return dataset_x, dataset_y, data

    def make_result_dataframe(self, xx_test, result1, result2, result3, original_data):
        """결과 데이터프레임 생성 (bid.ml.train.py와 동일)"""
        df_rst = pd.DataFrame(columns=self.result_columns)
        df_rst.astype(self.result_columns_type)
        
        selected_cols = ['입찰번호', '입찰차수', 
                         '기초금액', '낙찰하한률', '참여업체수', '간접비', '순공사원가', 
                         '면허제한코드', '공고기관코드', 
                         '키워드', '키워드점수', 
                         '공고기관명', '공고기관점수',
                         '공사지역', '공사지역점수',                             
                         '업체투찰률', '예가투찰률', '투찰률오차', 
                         '예정금액', '낙찰하한가', '낙찰금액']
        
        # 원본 데이터에서 필요한 컬럼들 선택
        df_test = original_data[selected_cols].copy()
        
        print("예측결과값 입력 시작")
        
        for i, v in enumerate(selected_cols):
            if v in df_test.columns:
                df_rst[v] = df_test[v]
                
        df_rst["업체투찰률예측"] = result1
        df_rst["예가투찰률예측"] = result2
        df_rst["참여업체수예측"] = result3
        
        # 예가투찰률 예측값으로부터 예정금액 역산 계산
        df_rst["예정금액예측"] = (df_rst["예가투찰률예측"] * df_rst["기초금액"]) / df_rst["낙찰하한률"]
        
        # 새로운 컬럼들 계산
        df_rst["낙찰금액(업체투찰률) 예측"] = df_rst["업체투찰률예측"] * df_rst["기초금액"]
        df_rst["A값여부"] = df_rst["업체투찰률예측"].apply(lambda x: 'O' if x >= 0.8 else '')
        
        # 결과1 계산
        def calculate_result1(row):
            predicted_amount = row["낙찰금액(업체투찰률) 예측"]
            min_bid = row["낙찰하한가"]
            actual_amount = row["낙찰금액"]
            
            if predicted_amount < min_bid:
                return "낙찰하한선미달"
            elif predicted_amount >= min_bid and predicted_amount < actual_amount:
                return "낙찰"
            else:
                return "-"
        
        df_rst["결과1"] = df_rst.apply(calculate_result1, axis=1)
        
        # 예정금액(예가투찰률) 예측
        df_rst["예정금액(예가투찰률) 예측"] = (df_rst["예가투찰률예측"] / df_rst["낙찰하한률"]) * df_rst["기초금액"]
        df_rst["예정금액*낙찰하한율"] = df_rst["예정금액(예가투찰률) 예측"] * df_rst["낙찰하한률"]
        
        # 결과2 계산
        def calculate_result2(row):
            predicted_amount = row["예정금액*낙찰하한율"]
            min_bid = row["낙찰하한가"]
            actual_amount = row["낙찰금액"]
            
            if predicted_amount < min_bid:
                return "낙찰하한선미달"
            elif predicted_amount >= min_bid and predicted_amount < actual_amount:
                return "낙찰"
            else:
                return "-"
        
        df_rst["결과2"] = df_rst.apply(calculate_result2, axis=1)
        
        print("예측결과테이블 작성완료")
        return df_rst

    def predict_and_evaluate(self, dataset_x, dataset_y):
        """예측 수행 및 성능 평가"""
        print("🤖 모델 예측 수행 중...")
        
        # 원본 학습 시 사용된 특성들만 선택 (bid.ml.train.py와 동일)
        # 인덱스 [0,1,2,7,8,13,17,19,21]에 해당하는 컬럼들만 사용
        original_features = ['기초금액', '낙찰하한률', '참여업체수', '간접비', '순공사원가', 
                           '면허제한코드', '공고기관점수', '공사지역점수', '키워드점수']
        
        # 존재하는 특성만 선택
        available_features = [col for col in original_features if col in dataset_x.columns]
        X_selected = dataset_x[available_features].fillna(0)
        
        print(f"📊 사용할 특성 수: {len(available_features)}개")
        print(f"📋 특성 목록: {available_features}")
        
        # 데이터 정규화
        X_scaled = self.scaler.transform(X_selected)
        
        # 예측 수행
        result1 = self.model1.predict(X_scaled)  # 업체투찰률
        result2 = self.model2.predict(X_scaled)  # 예가투찰률
        result3 = self.model3.predict(X_scaled)  # 참여업체수
        
        print("✅ 예측 완료")
        
        # 성능 평가
        print("\n📊 성능 평가:")
        print("="*60)
        
        # 업체투찰률 성능
        y1_actual = dataset_y['업체투찰률'].values
        mse1 = mean_squared_error(y1_actual, result1)
        r2_1 = r2_score(y1_actual, result1)
        mae1 = mean_absolute_error(y1_actual, result1)
        rmse1 = np.sqrt(mse1)
        
        print(f"업체투찰률 모델:")
        print(f"  R²:   {r2_1:.6f}")
        print(f"  RMSE: {rmse1:.6f}")
        print(f"  MAE:  {mae1:.6f}")
        
        # 예가투찰률 성능
        y2_actual = dataset_y['예가투찰률'].values
        mse2 = mean_squared_error(y2_actual, result2)
        r2_2 = r2_score(y2_actual, result2)
        mae2 = mean_absolute_error(y2_actual, result2)
        rmse2 = np.sqrt(mse2)
        
        print(f"\n예가투찰률 모델:")
        print(f"  R²:   {r2_2:.6f}")
        print(f"  RMSE: {rmse2:.6f}")
        print(f"  MAE:  {mae2:.6f}")
        
        # 참여업체수 성능
        y3_actual = dataset_y['참여업체수'].values
        mse3 = mean_squared_error(y3_actual, result3)
        r2_3 = r2_score(y3_actual, result3)
        mae3 = mean_absolute_error(y3_actual, result3)
        rmse3 = np.sqrt(mse3)
        
        print(f"\n참여업체수 모델:")
        print(f"  R²:   {r2_3:.6f}")
        print(f"  RMSE: {rmse3:.6f}")
        print(f"  MAE:  {mae3:.6f}")
        
        # 평균 성능
        avg_r2 = (r2_1 + r2_2 + r2_3) / 3
        avg_rmse = (rmse1 + rmse2 + rmse3) / 3
        
        print(f"\n평균 성능:")
        print(f"  평균 R²:   {avg_r2:.6f}")
        print(f"  평균 RMSE: {avg_rmse:.6f}")
        
        return result1, result2, result3

    def save_result_to_excel(self, df_result, filename):
        """결과를 엑셀 파일로 저장"""
        print(f"💾 엑셀 파일 저장 중: {filename}")
        
        try:
            df_result.to_excel(
                filename,
                sheet_name='Sheet1',
                na_rep='NaN',
                float_format="%.8f",
                header=True,
                index=True,
                index_label="id",
                startrow=1,
                startcol=1,
                freeze_panes=(2, 0)
            )
            
            # 통계 추가
            self.add_result_statistics(filename, df_result)
            
            print(f"✅ 엑셀 파일 저장 완료: {filename}")
            
        except Exception as e:
            print(f"❌ 엑셀 저장 실패: {e}")
            # CSV로 대체 저장
            csv_filename = filename.replace('.xlsx', '.csv')
            df_result.to_csv(csv_filename, 
                           na_rep='NaN', 
                           float_format="%.8f", 
                           header=True, 
                           index=True, 
                           index_label="id",
                           encoding='utf-8-sig')
            print(f"✅ CSV 파일로 저장: {csv_filename}")

    def add_result_statistics(self, filename, df_result):
        """엑셀 파일에 통계 추가"""
        try:
            if not OPENPYXL_AVAILABLE:
                return
                
            wb = load_workbook(filename)
            ws = wb.active
            
            data_rows = len(df_result)
            start_row = 3
            end_row = start_row + data_rows - 1
            
            # 결과1 통계
            stats_row = end_row + 2
            ws[f'AC{stats_row}'] = "=== 결과1 통계 ==="
            ws[f'AC{stats_row}'].font = openpyxl.styles.Font(bold=True, color="0000FF")
            
            ws[f'AC{stats_row + 1}'] = "낙찰 개수:"
            count_formula1 = f'=COUNTIF(AD{start_row}:AD{end_row},"낙찰")'
            ws[f'AD{stats_row + 1}'] = count_formula1
            
            ws[f'AC{stats_row + 2}'] = "낙찰 비율:"
            rate_formula1 = f'=IF(AD{stats_row + 1}>0,AD{stats_row + 1}/{data_rows}*100,0)'
            ws[f'AD{stats_row + 2}'] = rate_formula1
            
            # 결과2 통계
            ws[f'AF{stats_row}'] = "=== 결과2 통계 ==="
            ws[f'AF{stats_row}'].font = openpyxl.styles.Font(bold=True, color="00AA00")
            
            ws[f'AF{stats_row + 1}'] = "낙찰 개수:"
            count_formula2 = f'=COUNTIF(AG{start_row}:AG{end_row},"낙찰")'
            ws[f'AG{stats_row + 1}'] = count_formula2
            
            ws[f'AF{stats_row + 2}'] = "낙찰 비율:"
            rate_formula2 = f'=IF(AG{stats_row + 1}>0,AG{stats_row + 1}/{data_rows}*100,0)'
            ws[f'AG{stats_row + 2}'] = rate_formula2
            
            wb.save(filename)
            print("📊 통계 추가 완료")
            
        except Exception as e:
            print(f"⚠️ 통계 추가 실패: {e}")

    def test_model(self, test_filename):
        """전체 테스트 프로세스 실행"""
        print("🚀 개선된 모델 성능 테스트 시작")
        print("="*80)
        
        # 1. 테스트 데이터 로드 및 전처리
        dataset_x, dataset_y, original_data = self.load_test_data(test_filename)
        
        # 2. 예측 수행 및 성능 평가
        result1, result2, result3 = self.predict_and_evaluate(dataset_x, dataset_y)
        
        # 3. 결과 데이터프레임 생성
        print("\n📊 결과 데이터프레임 생성 중...")
        df_result = self.make_result_dataframe(dataset_x, result1, result2, result3, original_data)
        
        # 4. 엑셀 파일로 저장
        from datetime import datetime
        timestamp = datetime.now().strftime("%y%m%d%H%M")
        excel_filename = f"res/improved_model_test_result_{timestamp}.xlsx"
        self.save_result_to_excel(df_result, excel_filename)
        
        print(f"\n🎉 테스트 완료!")
        print(f"📁 결과 파일: {excel_filename}")
        print(f"📊 데이터 크기: {len(df_result)}행 x {len(df_result.columns)}열")

def main():
    """메인 함수"""
    print("🔍 개선된 모델 성능 테스트")
    print("="*80)
    
    # 테스터 초기화
    tester = ImprovedModelTester()
    
    # 테스트 실행 (원본 데이터로 테스트)
    tester.test_model('bid_250921_1.csv')

if __name__ == "__main__":
    main()
