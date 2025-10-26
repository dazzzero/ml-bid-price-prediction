# -*- coding: utf-8 -*-
"""
입찰 성공률 예측 시스템 - 머신러닝 모델 훈련 스크립트
Created on Mon Dec  2 10:14:33 2024

@author: user

이 파일의 목적:
- 조달청 입찰 데이터를 사용하여 머신러닝 모델을 훈련시키는 스크립트
- 3개의 MLPRegressor 모델을 훈련 (업체투찰률, 예가투찰률, 참여업체수 예측)
- 훈련된 모델과 전처리 도구들을 파일로 저장
- 텍스트 데이터를 TF-IDF 방식으로 벡터화하여 숫자로 변환

입찰 유형별 지원:
- 공사입찰: 순공사원가, 간접비, 주공종명 등 공사 전용 컬럼 포함
- 구매입찰: 기본 컬럼들만 사용 (공사 전용 컬럼 제외)
- 용역입찰: 기본 컬럼들만 사용 (공사 전용 컬럼 제외)

사용법:
- python bid.ml.train.py                    # 자동 감지 모드
- python bid.ml.train.py cst       # 공사입찰 모드
- python bid.ml.train.py mtrl          # 구매입찰 모드
- python bid.ml.train.py gdns           # 용역입찰 모드
- python bid.ml.train.py test              # 모델 성능 테스트
"""

# ===== 필요한 라이브러리들 import =====
import os  # 파일 경로 조작을 위한 라이브러리
import joblib  # 머신러닝 모델을 파일로 저장/불러오기 위한 라이브러리
import numpy as np  # 수치 계산을 위한 라이브러리 (행렬, 배열 연산)
import random as rnd  # 랜덤 숫자 생성

from time import time  # 실행 시간 측정을 위한 라이브러리
import pandas as pd  # 데이터 분석을 위한 라이브러리 (엑셀과 비슷한 기능)

# 엑셀 파일 처리를 위한 라이브러리 (통계 추가용)
try:
    import openpyxl
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("⚠️  openpyxl이 설치되지 않았습니다. 엑셀 통계 기능을 사용하려면 'pip install openpyxl'을 실행하세요.")

# 한국어 자연어 처리를 위한 라이브러리
from kiwipiepy import Kiwi  # 한국어 형태소 분석기 (단어를 쪼개는 도구)

# 텍스트를 숫자로 변환하는 도구들
from sklearn.feature_extraction.text import TfidfVectorizer  # 텍스트를 숫자로 변환
from scipy.sparse import csr_matrix  # 메모리 효율적인 행렬 저장 방식

# 머신러닝 모델과 전처리 도구들
from sklearn.neural_network import MLPRegressor  # 인공신경망 회귀 모델 (뇌의 뉴런처럼 작동)
from sklearn.preprocessing import StandardScaler  # 데이터를 정규화하는 도구 (0~1 사이로 맞춤)

# 데이터 분할 도구
from sklearn.model_selection import train_test_split  # 훈련 데이터와 테스트 데이터로 분할

# 고급 특성 엔지니어링
from advanced_feature_engineering import AdvancedFeatureEngineering


class KiwiTokenizer():
    """
    한국어 텍스트를 처리하는 클래스 (predict.py와 동일)
    
    이 클래스의 역할:
    1. 한국어 문장을 단어 단위로 쪼개기 (형태소 분석)
    2. 의미있는 단어만 추출하기 (명사, 형용사 등)
    3. 불필요한 문자 제거하기 (괄호, 특수문자 등)
    
    예시: "서울시청 건물 신축공사" → ["서울시청", "건물", "신축", "공사"]
    """
    
    def __init__(self, saved_filenm):
        """
        KiwiTokenizer 초기화 함수
        
        Args:
            saved_filenm (str): 저장된 사전 파일명 (이전에 학습한 단어사전을 불러올 때 사용)
        """
        # 랜덤 번호 생성 (파일명에 사용)
        self.rnd_num = rnd.randint(100, 999)
        
        # 현재 작업 디렉토리 경로 설정
        self.cur_dir = os.getcwd()  # 현재 폴더 경로 가져오기
        self.data_dir = self.cur_dir+'\\data\\'  # 데이터 폴더 경로 (입찰 데이터가 있는 곳)
        self.save_dir = self.cur_dir+'\\res\\'  # 결과 저장 폴더 (모델 파일들이 있는 곳)
        self.save_filename = saved_filenm  # 저장할 파일명
        
        # Kiwi 형태소 분석기 초기화
        self.kiwi = None
        self.kiwi = self.CreateKiwi(saved_filenm)  # Kiwi 객체 생성
    
    def save(self, filename):
        joblib.dump(self.kiwi._user_values, self.save_dir+filename)
        return self
    
    def load(self, filename):
        o = joblib.load(self.save_dir+filename)
        return o
        
    def loadDictonary(self, filename):
        self.data = pd.read_csv(self.data_dir+filename)   #표준국어대사전.NNP.csv
        print(self.data[:10])
        for i, row in enumerate(self.data.iloc):
            #print(i, row["단어"])
            self.kiwi.add_user_word(row["단어"], 'NNP')
    
    def CreateKiwi(self, saved_filenm):
        o = None
        if(self.kiwi is None):
            if(saved_filenm is None):
                o = Kiwi(num_workers=8)
            else:
                o = Kiwi(num_workers=8)
                o._user_values = self.load(self.save_dir+saved_filenm)
                #o.load_user_dictionary(self.save_dir+saved_filenm)
        
        return o
            
    def cleared_line(self, line):
        # numpy.float64나 NaN 값 처리
        if pd.isna(line) or line is None:
            line = ''
        else:
            line = str(line).lower().replace('(', ' ').replace(')', ' ').replace('n/a', '')
        
        nm_words = []
        tokens = self.kiwi.tokenize(line)
        for token in tokens:
            if token.tag in ['MM', 'NNG', 'NNB', 'NNP', 'SL', 'XPN', 'MAG', 'SN', 'SO', 'W_SERIAL']:
                nm_words.append(token.form)
                
        return ' '.join(nm_words)
    
    def cleared_lines_from(self, orglines):
        lines = []
        for line in orglines:
            lines.append(self.cleared_line(line))
        
        return lines      
    
    def get_key(self, voca, val):
      
        for key, value in voca.items():
            if val == value:
                return key
    
        return "key doesn't exist"
    
    def conv_words(self, voca, vals):
        ret = []
        for val in vals:
            for key, value in voca.items():
                if val == value:
                    ret.append(key)
            
            
        return " ".join(ret)
    
    
    def nn_only(self, orglines):
        lines = []
        for key in orglines:
            # numpy.float64나 NaN 값 처리
            if pd.isna(key) or key is None:
                key = ''
            else:
                key = str(key).lower().replace('(', ' ').replace(')', ' ').replace('n/a', '')
            
            nm_words = []
            tokens = self.kiwi.tokenize(key)
            #print(key, tokens)
            for token in tokens:
                if token.tag in ['MM', 'NNG', 'NNB', 'NNP', 'SL', 'XPN', 'MAG', 'SN', 'SO', 'W_SERIAL']:
                    nm_words.append(token.form)
            lines.append(' '.join(nm_words))
            
        return lines


class KiwiVectorizer():
    def __init__(self):
        self.rnd_num = rnd.randint(100, 999)
        
        self.cur_dir = os.getcwd()
        self.save_dir = self.cur_dir+'\\res\\'
        
        # 메모리 효율성을 위해 파라미터 최적화
        self.vect = TfidfVectorizer(
            ngram_range=(1, 1), 
            token_pattern='(?u)\\b\\w+\\b',
            max_features=5000,  # 최대 특성 수 제한
            min_df=2,           # 최소 문서 빈도 (2회 이상 나타나는 단어만)
            max_df=0.95,        # 최대 문서 빈도 (95% 이상 문서에 나타나는 단어 제외)
            sublinear_tf=True   # TF 스케일링으로 메모리 절약
        )
    
    def load(self, filename):
        voca = joblib.load(self.save_dir + filename)
        self.vect.vocabulary_ = voca["vocabulary"] 
        self.vect.idf_ = voca["idf"]
    
    def save(self, filename):
        joblib.dump({'vocabulary':self.vect.vocabulary_, 'idf':self.vect.idf_}, self.save_dir + filename)
        print("단어사전 파일("+filename+")로 저장합니다. ")
    
    def fit(self, lines):
        return self.vect.fit(lines)
    
    def transform(self, lines):
        # 메모리 효율성을 위해 toarray() 없이 CSR 행렬 그대로 반환
        return self.vect.transform(lines)
    
    def scores(self, lines):
        return self.toValues(self.transform(lines))


    def toValues(self, csrmat:csr_matrix):
        lst = []
        safty_sz = len(csrmat.indptr)-1
        for i, n in enumerate(csrmat.indptr):
            sumval = 0
            if i<safty_sz:
                indiceVal = csrmat.indices[csrmat.indptr[i]:csrmat.indptr[i+1]]
                dataVal = csrmat.data[csrmat.indptr[i]:csrmat.indptr[i+1]]
                for j, va1 in enumerate(indiceVal):
                    sumval += va1 * dataVal[j]
                
                lst.append(sumval)
        return lst


class BidLowerMarginRateTrain():
    """
    입찰 투찰률 예측 모델을 훈련시키는 메인 클래스
    
    이 클래스의 역할:
    1. 조달청 입찰 데이터를 불러와서 전처리
    2. 텍스트 데이터(키워드, 기관명, 지역)를 숫자로 변환
    3. 3개의 MLPRegressor 모델을 훈련
       - 모델1: 업체 투찰률 예측
       - 모델2: 예가 투찰률 예측  
       - 모델3: 참여 업체 수 예측
    4. 훈련된 모델과 전처리 도구들을 파일로 저장
    5. 예측 결과를 엑셀 파일로 출력
    
    머신러닝 과정:
    1. 데이터 로드 → 2. 텍스트 벡터화 → 3. 데이터 정규화 → 4. 모델 훈련 → 5. 모델 저장
    """
    
    def __init__(self, bid_type='auto'):
        """
        BidLowerMarginRateTrain 초기화 함수
        - 디렉토리 경로 설정
        - 데이터 컬럼 정의 (입찰 유형별)
        - 텍스트 처리 도구들 초기화
        
        Args:
            bid_type (str): 입찰 유형 ('cst', 'mtrl', 'gdns', 'auto')
                - 'cst': 공사입찰 (기존 컬럼들 + 순공사원가, 간접비 등)
                - 'mtrl': 구매입찰 (기본 컬럼들만)
                - 'gdns': 용역입찰 (기본 컬럼들만)
                - 'auto': 자동 감지 (데이터 로드 시 컬럼 존재 여부로 판단)
        """
        # 랜덤 번호 생성 (파일명에 사용)
        self.rnd_num = rnd.randint(100, 999)
        
        # 디렉토리 경로 설정
        self.cur_dir = os.getcwd()  # 현재 작업 디렉토리
        self.data_dir = self.cur_dir+'\\data\\'  # 데이터 폴더 (입찰 데이터가 있는 곳)
        self.save_dir = self.cur_dir+'\\res\\'  # 결과 저장 폴더 (모델 파일들이 저장될 곳)
        
        # 결과 엑셀 파일 경로 설정 (나중에 데이터 크기와 테스트 비율에 따라 생성)
        self.excel_file_nm = None  # 결과 엑셀 파일명 (나중에 설정)
        self.xlxs_dir = None  # 엑셀 파일 전체 경로 (나중에 설정) 
        
        # 입찰 유형 저장
        self.bid_type = bid_type
        
        # ===== 입찰 유형별 컬럼 정의 =====
        self._define_bid_type_columns()
        
        # ===== 텍스트 처리 도구들 초기화 =====
        # 한국어 형태소 분석기 초기화 (새로 생성)
        self.tokenizer = KiwiTokenizer(None)
        self.tokenizer.loadDictonary('표준국어대사전.NNP.csv')  # 표준국어대사전을 로드하여 정확한 단어 인식
        
        # TF-IDF 벡터화기 초기화 (텍스트를 숫자로 변환하는 도구)
        self.vectorizer = KiwiVectorizer()
        
        # ===== 결과 데이터 컬럼 정의 =====
        # 예측 결과를 저장할 엑셀 파일의 컬럼들 정의 (원본 데이터 + 예측 결과)
        self.result_columns = self._get_result_columns()
        
        # 결과 컬럼들의 데이터 타입 정의 (동적으로 생성)
        self.result_columns_type = self._get_result_column_types()

    def _define_bid_type_columns(self):
        """
        입찰 유형별로 사용할 컬럼들을 정의하는 함수
        
        입찰 유형별 차이점:
        - 공사입찰: 순공사원가, 간접비, A계산여부, 순공사원가적용여부, 주공종명 포함
        - 구매입찰/용역입찰: 위 컬럼들 제외하고 기본 컬럼들만 사용
        """
        # 기본 컬럼들 (모든 입찰 유형에서 공통으로 사용)
        base_columns = [
            '기초금액', '낙찰하한률', '참여업체수', 
            '낙찰금액', '업체투찰률', '예가투찰률', '투찰률오차', 
            '입찰번호', '입찰차수', 
            '예정금액', '낙찰하한가', 
            '면허제한코드', '공고기관코드',
            '공고기관명', '공고기관점수',
            '공사지역', '공사지역점수',
            '키워드', '키워드점수'
        ]
        
        # 공사입찰 전용 컬럼들
        construction_only_columns = [
            '간접비', '순공사원가', '주공종명'
        ]
        
        # 입찰 유형별 컬럼 정의
        if self.bid_type == 'cst':
            # 공사입찰: 기본 컬럼 + 공사입찰 전용 컬럼
            self.cvs_columns = base_columns + construction_only_columns
            self.bid_type_name = "공사입찰"
        elif self.bid_type in ['mtrl', 'gdns']:
            # 구매입찰/용역입찰: 기본 컬럼만 사용
            self.cvs_columns = base_columns.copy()
            if self.bid_type == 'mtrl':
                self.bid_type_name = "구매입찰"
            else:
                self.bid_type_name = "용역입찰"
        else:  # 'auto' 또는 기타
            # 자동 감지: 기본 컬럼으로 시작하고, 데이터 로드 시 동적으로 조정
            self.cvs_columns = base_columns.copy()
            self.bid_type_name = "자동감지"
        
        # 컬럼 데이터 타입 정의 (동적으로 생성)
        self.cvs_columns_type = self._get_column_types()

    def _get_column_types(self):
        """
        현재 정의된 컬럼들에 대한 데이터 타입을 반환하는 함수
        
        Returns:
            dict: 컬럼명을 키로, 데이터 타입을 값으로 하는 딕셔너리
        """
        # 기본 데이터 타입 정의
        type_mapping = {
            # 숫자형 컬럼들
            '기초금액': 'float64',
            '낙찰하한률': 'float64', 
            '참여업체수': 'float64',
            '낙찰금액': 'int64',
            '업체투찰률': 'float64',
            '예가투찰률': 'float64',
            '투찰률오차': 'float64',
            '간접비': 'int64',
            '순공사원가': 'int64',
            '입찰차수': 'int64',
            '예정금액': 'int64',
            '낙찰하한가': 'int64',
            '공고기관점수': 'float64',
            '공사지역점수': 'float64',
            '키워드점수': 'float64',
            
            # 문자열 컬럼들
            '입찰번호': 'str',
            '면허제한코드': 'str',
            '공고기관코드': 'str',
            '주공종명': 'str',
            '공고기관명': 'str',
            '공사지역': 'str',
            '키워드': 'str'
        }
        
        # 현재 정의된 컬럼들에 대해서만 타입 반환
        return {col: type_mapping.get(col, 'str') for col in self.cvs_columns}

    def _get_result_columns(self):
        """
        결과 데이터프레임에 사용할 컬럼들을 동적으로 생성하는 함수
        
        Returns:
            list: 결과 컬럼 리스트
        """
        # 기본 결과 컬럼들 (입찰 유형에 관계없이 공통)
        base_result_columns = [
            '입찰번호', '입찰차수', '기초금액', 
            '낙찰하한률', '참여업체수', 
            '면허제한코드', '공고기관코드',
            '공고기관명', '공고기관점수',
            '공사지역', '공사지역점수',                               
            '키워드', '키워드점수',
            '업체투찰률', '예가투찰률', '투찰률오차', 
            '예정금액', '낙찰하한가', '낙찰금액', 
            '업체투찰률예측', '예가투찰률예측', '참여업체수예측', '예정금액예측',
            '낙찰금액(업체투찰률) 예측', 'A값여부', '결과1', 
            '예정금액(예가투찰률) 예측', '예정금액*낙찰하한율', '결과2'
        ]
        
        # 공사입찰 전용 컬럼들
        construction_result_columns = ['간접비', '순공사원가']
        
        # 입찰 유형에 따라 컬럼 결정
        if self.bid_type == 'cst':
            # 공사입찰: 기본 컬럼 + 공사입찰 전용 컬럼
            result_columns = base_result_columns.copy()
            # 간접비와 순공사원가를 적절한 위치에 삽입
            insert_index = result_columns.index('참여업체수') + 1
            result_columns[insert_index:insert_index] = construction_result_columns
        else:
            # 구매입찰/용역입찰: 기본 컬럼만 사용
            result_columns = base_result_columns.copy()
        
        return result_columns

    def _get_result_column_types(self):
        """
        결과 컬럼들의 데이터 타입을 동적으로 생성하는 함수
        
        Returns:
            dict: 결과 컬럼명을 키로, 데이터 타입을 값으로 하는 딕셔너리
        """
        # 기본 결과 컬럼 타입 정의
        result_type_mapping = {
            # 기본 데이터 컬럼들
            '입찰번호': 'str',
            '입찰차수': 'int64',
            '기초금액': 'float64',
            '낙찰하한률': 'float64',
            '참여업체수': 'float64',
            '간접비': 'int64',
            '순공사원가': 'int64',
            '면허제한코드': 'float64',
            '공고기관코드': 'float64',
            '공고기관명': 'str',
            '공고기관점수': 'float64',
            '공사지역': 'str',
            '공사지역점수': 'float64',
            '키워드': 'str',
            '키워드점수': 'float64',
            
            # 실제값 컬럼들
            '업체투찰률': 'float64',
            '예가투찰률': 'float64',
            '투찰률오차': 'float64',
            '예정금액': 'int64',
            '낙찰하한가': 'int64',
            '낙찰금액': 'int64',
            
            # 예측값 컬럼들
            '업체투찰률예측': 'float64',
            '예가투찰률예측': 'float64',
            '참여업체수예측': 'float64',
            '예정금액예측': 'int64',
            '낙찰금액(업체투찰률) 예측': 'float64',
            'A값여부': 'str',
            '결과1': 'str',
            '예정금액(예가투찰률) 예측': 'float64',
            '예정금액*낙찰하한율': 'float64',
            '결과2': 'str'
        }
        
        # 현재 정의된 결과 컬럼들에 대해서만 타입 반환
        return {col: result_type_mapping.get(col, 'str') for col in self.result_columns}

    def _detect_bid_type(self, data):
        """
        데이터의 컬럼 존재 여부를 기반으로 입찰 유형을 자동 감지하는 함수
        
        Args:
            data (pandas.DataFrame): 로드된 데이터
            
        감지 로직:
        - 간접비, 순공사원가, 주공종명이 모두 있으면 → 공사입찰
        - 위 컬럼들이 없으면 → 구매입찰 또는 용역입찰 (기본값: 구매입찰)
        """
        print("="*80)
        print("🔍 입찰 유형 자동 감지 중...")
        
        # 공사입찰 전용 컬럼들 확인
        construction_columns = ['간접비', '순공사원가', '주공종명']
        available_construction_columns = [col for col in construction_columns if col in data.columns]
        
        print(f"공사입찰 전용 컬럼 확인:")
        print(f"  - 간접비: {'있음' if '간접비' in data.columns else '없음'}")
        print(f"  - 순공사원가: {'있음' if '순공사원가' in data.columns else '없음'}")
        print(f"  - 주공종명: {'있음' if '주공종명' in data.columns else '없음'}")
        
        # 입찰 유형 결정
        if len(available_construction_columns) >= 2:  # 2개 이상의 공사입찰 컬럼이 있으면
            detected_type = 'cst'
            self.bid_type_name = "공사입찰"
            print(f"✅ 감지된 입찰 유형: {self.bid_type_name}")
        else:
            detected_type = 'mtrl'  # 기본값을 구매입찰로 설정
            self.bid_type_name = "구매입찰"
            print(f"✅ 감지된 입찰 유형: {self.bid_type_name} (기본값)")
        
        # 컬럼 정의 업데이트
        self.bid_type = detected_type
        self._define_bid_type_columns()
        
        print(f"📊 사용할 컬럼 수: {len(self.cvs_columns)}개")
        print(f"📋 컬럼 목록: {self.cvs_columns}")
        print("="*80)

    def _prepare_dataset_x(self, data):
        """
        데이터에서 필요한 컬럼들을 추출하고 누락된 컬럼은 기본값으로 처리하는 함수
        
        Args:
            data (pandas.DataFrame): 원본 데이터
            
        Returns:
            pandas.DataFrame: 처리된 데이터프레임
        """
        print("="*80)
        print("📊 데이터셋 준비 중...")
        
        # 필요한 컬럼들 확인
        available_columns = [col for col in self.cvs_columns if col in data.columns]
        missing_columns = [col for col in self.cvs_columns if col not in data.columns]
        
        print(f"✅ 사용 가능한 컬럼: {len(available_columns)}개")
        print(f"⚠️  누락된 컬럼: {len(missing_columns)}개")
        
        if missing_columns:
            print(f"누락된 컬럼 목록: {missing_columns}")
        
        # 사용 가능한 컬럼들로 데이터프레임 생성
        dataset_x = pd.DataFrame(data, columns=available_columns)
        
        # 누락된 컬럼들을 기본값으로 추가
        for col in missing_columns:
            if col in ['간접비', '순공사원가']:
                # 숫자형 컬럼은 0으로 채움
                dataset_x[col] = 0
                print(f"  - {col}: 기본값 0으로 설정")
            elif col == '주공종명':
                # 문자열 컬럼은 빈 문자열로 채움
                dataset_x[col] = ''
                print(f"  - {col}: 기본값 빈 문자열로 설정")
            else:
                # 기타 컬럼은 기본값으로 설정
                dataset_x[col] = 0 if col in ['공고기관점수', '공사지역점수', '키워드점수'] else ''
                print(f"  - {col}: 기본값으로 설정")
        
        # 컬럼 순서 정렬 (정의된 순서대로)
        dataset_x = dataset_x[self.cvs_columns]
        
        print(f"✅ 최종 데이터셋 크기: {dataset_x.shape}")
        print("="*80)
        
        return dataset_x

    def _get_selected_column_indices(self):
        """
        입찰 유형에 따라 사용할 컬럼의 인덱스를 반환하는 함수
        
        Returns:
            list: 선택된 컬럼의 인덱스 리스트
        """
        # 기본적으로 사용할 컬럼들 (모든 입찰 유형에서 공통)
        base_columns = ['기초금액', '낙찰하한률', '참여업체수', '면허제한코드', '공고기관점수', '공사지역점수', '키워드점수']
        
        # 공사입찰 전용 컬럼들
        construction_columns = ['간접비', '순공사원가']
        
        # 입찰 유형에 따라 컬럼 결정
        if self.bid_type == 'cst':
            # 공사입찰: 기본 컬럼 + 공사입찰 전용 컬럼
            selected_columns = base_columns[:3] + construction_columns + base_columns[3:]
        else:
            # 구매입찰/용역입찰: 기본 컬럼만 사용
            selected_columns = base_columns
        
        # 컬럼명을 인덱스로 변환
        column_indices = []
        for col in selected_columns:
            if col in self.cvs_columns:
                column_indices.append(self.cvs_columns.index(col))
            else:
                print(f"⚠️  경고: 컬럼 '{col}'을 찾을 수 없습니다.")
        
        print(f"선택된 컬럼들: {selected_columns}")
        print(f"해당 인덱스들: {column_indices}")
        
        return column_indices

    def _get_result_selected_columns(self):
        """
        결과 데이터프레임에서 사용할 컬럼들을 동적으로 반환하는 함수
        
        Returns:
            list: 선택된 컬럼 리스트
        """
        # 기본 컬럼들 (모든 입찰 유형에서 공통)
        base_cols = [
            '입찰번호', '입찰차수', 
            '기초금액', '낙찰하한률', '참여업체수', 
            '면허제한코드', '공고기관코드', 
            '키워드', '키워드점수', 
            '공고기관명', '공고기관점수',
            '공사지역', '공사지역점수',                             
            '업체투찰률', '예가투찰률', '투찰률오차', 
            '예정금액', '낙찰하한가', '낙찰금액'
        ]
        
        # 공사입찰 전용 컬럼들
        construction_cols = ['간접비', '순공사원가']
        
        # 입찰 유형에 따라 컬럼 결정
        if self.bid_type == 'cst':
            # 공사입찰: 기본 컬럼 + 공사입찰 전용 컬럼
            selected_cols = base_cols[:5] + construction_cols + base_cols[5:]
        else:
            # 구매입찰/용역입찰: 기본 컬럼만 사용
            selected_cols = base_cols
        
        # 실제로 존재하는 컬럼들만 필터링
        available_cols = [col for col in selected_cols if col in self.cvs_columns]
        
        print(f"결과용 선택된 컬럼들: {available_cols}")
        
        return available_cols
        
        # ===== 텍스트 처리 도구들 초기화 =====
        # 한국어 형태소 분석기 초기화 (새로 생성)
        self.tokenizer = KiwiTokenizer(None)
        self.tokenizer.loadDictonary('표준국어대사전.NNP.csv')  # 표준국어대사전을 로드하여 정확한 단어 인식
        
        # TF-IDF 벡터화기 초기화 (텍스트를 숫자로 변환하는 도구)
        self.vectorizer = KiwiVectorizer()

    def generateExcelFileName(self, data_size, test_ratio):
        """
        데이터 크기와 테스트 비율에 따라 엑셀 파일명을 생성하는 함수
        
        Args:
            data_size (int): 전체 데이터 크기
            test_ratio (float): 테스트 데이터 비율 (0.0 ~ 1.0)
            
        Returns:
            str: 생성된 파일명 (result-yyMMddMMss-데이터갯수-테스트비율.xlsx)
            
        파일명 규칙:
        - result-yyMMddMMss-데이터갯수-테스트비율.xlsx
        - 예: result-2412011430-5-8515.xlsx (2024년 12월 1일 14시 30분, 5만개 데이터, 85:15 비율)
        """
        from datetime import datetime
        
        # 현재 시간을 yyMMddMMss 형식으로 변환
        now = datetime.now()
        time_str = now.strftime("%y%m%d%H%M")
        
        # 데이터 크기를 만개 단위로 변환
        if data_size < 50000:
            data_unit = 1  # 5만개 미만은 1로 표시
        elif data_size < 100000:
            data_unit = 5  # 5만~10만개는 5로 표시
        elif data_size < 300000:
            data_unit = 10  # 10만~30만개는 10으로 표시
        else:
            data_unit = 30  # 30만개 이상은 30으로 표시
        
        # 테스트 비율을 정수로 변환 (예: 0.15 -> 15, 0.2 -> 20)
        train_ratio = int((1 - test_ratio) * 100)
        test_ratio_int = int(test_ratio * 100)
        ratio_str = f"{train_ratio}{test_ratio_int:02d}"  # 8515, 8020 등
        
        # 파일명 생성
        filename = f"result-{time_str}-{data_unit}-{ratio_str}.xlsx"
        
        return filename

    
    def make_result_dataframe2(self, xx_test, result1, result2, result3):
        df_rst = pd.DataFrame(columns = self.result_columns)
        df_rst.astype(self.result_columns_type)        
        
        # 입찰 유형에 따라 동적으로 컬럼 선택
        selected_cols = self._get_result_selected_columns()
        
        df_xx_test = pd.DataFrame(xx_test, columns = self.cvs_columns)
        df_test = pd.DataFrame(df_xx_test, columns = selected_cols)
        
        print("예측결과값 입력 시작")
        
        for i, v in enumerate(selected_cols):
            df_rst[v] = df_test[v]
                
        df_rst["업체투찰률예측"] = result1
        df_rst["예가투찰률예측"] = result2
        df_rst["참여업체수예측"] = result3
        
        # 예가투찰률 예측값으로부터 예정금액 역산 계산
        # 예가투찰률 = (예정금액 / 기초금액) * 낙찰하한률
        # 따라서 예정금액 = (예가투찰률 * 기초금액) / 낙찰하한률
        df_rst["예정금액예측"] = (df_rst["예가투찰률예측"] * df_rst["기초금액"]) / df_rst["낙찰하한률"]
        
        # ===== 새로운 컬럼들 계산 =====
        # 낙찰금액(업체투찰률) 예측 = 업체투찰률예측 * 기초금액
        df_rst["낙찰금액(업체투찰률) 예측"] = df_rst["업체투찰률예측"] * df_rst["기초금액"]
        
        # A값여부: 업체투찰률예측이 0.8 이상이면 'O', 아니면 빈 문자열
        df_rst["A값여부"] = df_rst["업체투찰률예측"].apply(lambda x: 'O' if x >= 0.8 else '')
        
        # 결과1: 낙찰금액(업체투찰률) 예측이 낙찰하한가보다 작으면 "낙찰하한선미달", 
        # 낙찰하한가 이상이고 낙찰금액 미만이면 "낙찰", 아니면 "-"
        def calculate_result1(row):
            predicted_amount = row["낙찰금액(업체투찰률) 예측"]
            min_bid = row["낙찰하한가"]
            actual_amount = row["낙찰금액"]
            
            if predicted_amount < min_bid:
                return "낙찰하한선미달"
            elif predicted_amount >= min_bid and predicted_amount < actual_amount:
                return "낙찰"
            else:
                return "-"
        
        df_rst["결과1"] = df_rst.apply(calculate_result1, axis=1)
        
        # 예정금액(예가투찰률) 예측 = 예가투찰률예측 / 낙찰하한률 * 기초금액
        df_rst["예정금액(예가투찰률) 예측"] = (df_rst["예가투찰률예측"] / df_rst["낙찰하한률"]) * df_rst["기초금액"]
        
        # 예정금액*낙찰하한율 = 예정금액(예가투찰률) 예측 * 낙찰하한률
        df_rst["예정금액*낙찰하한율"] = df_rst["예정금액(예가투찰률) 예측"] * df_rst["낙찰하한률"]
        
        # 결과2: 예정금액*낙찰하한율이 낙찰하한가보다 작으면 "낙찰하한선미달",
        # 낙찰하한가 이상이고 낙찰금액 미만이면 "낙찰", 아니면 "-"
        def calculate_result2(row):
            predicted_amount = row["예정금액*낙찰하한율"]
            min_bid = row["낙찰하한가"]
            actual_amount = row["낙찰금액"]
            
            if predicted_amount < min_bid:
                return "낙찰하한선미달"
            elif predicted_amount >= min_bid and predicted_amount < actual_amount:
                return "낙찰"
            else:
                return "-"
        
        df_rst["결과2"] = df_rst.apply(calculate_result2, axis=1)
        
        print("예측결과테이블 작성완료")
         
        return df_rst


    
    def arrayToDataFrame(self, data, cols):
        df = pd.DataFrame(pd.DataFrame(data), columns=cols)
        return df
    

    def loadTrainsetFromFile(self, filename):
        """
        CSV 파일에서 입찰 데이터를 불러와서 전처리하는 함수
        
        Args:
            filename (str): 불러올 CSV 파일명
            
        Returns:
            tuple: (x_train, x_test, y_train, y_test) - 훈련/테스트 데이터
            
        처리 과정:
        1. CSV 파일 읽기
        2. 입찰 유형 자동 감지 (auto 모드인 경우)
        3. 텍스트 데이터(키워드, 기관명, 지역)를 TF-IDF 점수로 변환
        4. 훈련 데이터와 테스트 데이터로 분할
        5. 필요한 컬럼만 선택하여 반환
        """
        print("원시 훈련데이타를 불러옵니다.")
        print(f"입찰 유형: {self.bid_type_name}")
        
        # CSV 파일 읽기
        data = pd.read_csv(self.data_dir+filename)
        
        # 입찰 유형 자동 감지 (auto 모드인 경우)
        if self.bid_type == 'auto':
            self._detect_bid_type(data)
        
        
        print("총 데이타수: "+str(len(data))+'건')
        
        print("="*80)
        print("학습셋과 테스트셋 분리")
        print(f"전체 데이터: {len(data)}개")
        
        # ===== 데이터 크기에 따른 동적 테스트 비율 설정 =====
        data_size = len(data)
        if data_size < 20000:
            test_ratio = 0.2  # 20% - 2만개 미만
            ratio_desc = "80:20 (소규모 데이터)"
        elif data_size < 50000:
            test_ratio = 0.15  # 15% - 2만~5만개
            ratio_desc = "85:15 (중소규모 데이터)"
        elif data_size < 100000:
            test_ratio = 0.1   # 10% - 5만~10만개
            ratio_desc = "90:10 (중규모 데이터)"
        else:
            test_ratio = 0.05  # 5% - 10만개 이상
            ratio_desc = "95:5 (대규모 데이터)"
        
        train_count = int(data_size * (1 - test_ratio))
        test_count = data_size - train_count
        
        print(f"적용 비율: {ratio_desc}")
        print(f"훈련 데이터: {train_count:,}개 ({((1-test_ratio)*100):.0f}%)")
        print(f"테스트 데이터: {test_count:,}개 ({test_ratio*100:.0f}%)")
        print("="*80)
        
        # ===== 입력 데이터(X) 준비 =====
        # 필요한 컬럼들만 선택하여 데이터프레임 생성 (누락된 컬럼은 기본값으로 처리)
        dataset_x = self._prepare_dataset_x(data)
        
        # 텍스트 컬럼들을 먼저 문자열로 변환
        print("텍스트 컬럼을 문자열로 변환 중...")
        if '키워드' in dataset_x.columns:
            dataset_x['키워드'] = dataset_x['키워드'].fillna('').astype(str)
            dataset_x['키워드'] = dataset_x['키워드'].replace(['nan', 'NaN', 'None', 'null'], '')
        if '공고기관명' in dataset_x.columns:
            dataset_x['공고기관명'] = dataset_x['공고기관명'].fillna('').astype(str)
            dataset_x['공고기관명'] = dataset_x['공고기관명'].replace(['nan', 'NaN', 'None', 'null'], '')
        if '공사지역' in dataset_x.columns:
            dataset_x['공사지역'] = dataset_x['공사지역'].fillna('').astype(str)
            dataset_x['공사지역'] = dataset_x['공사지역'].replace(['nan', 'NaN', 'None', 'null'], '')
        
        # 문자열 컬럼들을 숫자로 변환
        print("문자열 컬럼을 숫자로 변환 중...")
        
        # 면허제한코드를 숫자로 변환 (문자열을 해시값으로 변환)
        if '면허제한코드' in dataset_x.columns:
            dataset_x['면허제한코드'] = dataset_x['면허제한코드'].astype(str).apply(lambda x: hash(x) % 1000000 if pd.notna(x) and x != 'nan' else 0)
        
        # 공고기관코드를 숫자로 변환
        if '공고기관코드' in dataset_x.columns:
            dataset_x['공고기관코드'] = dataset_x['공고기관코드'].astype(str).apply(lambda x: hash(x) % 1000000 if pd.notna(x) and x != 'nan' else 0)
        
        # 각 컬럼의 데이터 타입을 올바르게 설정
        dataset_x.astype(self.cvs_columns_type)
        
        
        # ===== 텍스트 데이터 처리 =====
        # 각 텍스트 컬럼을 형태소 분석하여 의미있는 단어만 추출
        print("키워드 컬럼 데이터 타입 및 샘플:")
        print(f"키워드 컬럼 타입: {dataset_x['키워드'].dtype}")
        print(f"키워드 샘플 (처음 10개): {dataset_x['키워드'].head(10).tolist()}")
        print(f"키워드 NaN 개수: {dataset_x['키워드'].isna().sum()}")
        
        print("\n공고기관명 컬럼 데이터 타입 및 샘플:")
        print(f"공고기관명 컬럼 타입: {dataset_x['공고기관명'].dtype}")
        print(f"공고기관명 샘플 (처음 10개): {dataset_x['공고기관명'].head(10).tolist()}")
        print(f"공고기관명 NaN 개수: {dataset_x['공고기관명'].isna().sum()}")
        print(f"공고기관명 고유값 개수: {dataset_x['공고기관명'].nunique()}")
        
        print("\n공사지역 컬럼 데이터 타입 및 샘플:")
        print(f"공사지역 컬럼 타입: {dataset_x['공사지역'].dtype}")
        print(f"공사지역 샘플 (처음 10개): {dataset_x['공사지역'].head(10).tolist()}")
        print(f"공사지역 NaN 개수: {dataset_x['공사지역'].isna().sum()}")
        
        lines = self.tokenizer.nn_only(np.squeeze(dataset_x["키워드"].tolist()))      # 키워드 처리
        lines2 = self.tokenizer.nn_only(np.squeeze(dataset_x["공고기관명"].tolist()))  # 공고기관명 처리
        lines3 = self.tokenizer.nn_only(np.squeeze(dataset_x["공사지역"].tolist()))    # 공사지역 처리
        
        # 처리 결과 미리보기 (처음 10개)
        print("키워드 처리 결과:")
        print(lines[:10])
        print("공고기관명 처리 결과:")
        print(lines2[:10])
        print("공사지역 처리 결과:")
        print(lines3[:10])
        
        # ===== TF-IDF 벡터화기 학습 =====
        # 모든 텍스트 데이터를 합쳐서 단어사전을 학습
        all_text = lines + lines2 + lines3
        print(f"전체 텍스트 샘플 수: {len(all_text)}")
        print(f"빈 문자열 개수: {sum(1 for text in all_text if text.strip() == '')}")
        
        # 빈 문자열이 아닌 텍스트만 사용하여 학습
        non_empty_text = [text for text in all_text if text.strip() != '']
        print(f"비어있지 않은 텍스트 수: {len(non_empty_text)}")
        
        if len(non_empty_text) > 0:
            self.vectorizer.fit(non_empty_text)
        else:
            # 모든 텍스트가 비어있다면 기본값으로 학습
            self.vectorizer.fit(['기본키워드'])
        
        # ===== 텍스트를 TF-IDF 점수로 변환 =====
        pts = self.vectorizer.scores(lines)      # 키워드 점수 계산
        pts2 = self.vectorizer.scores(lines2)    # 공고기관명 점수 계산
        pts3 = self.vectorizer.scores(lines3)    # 공사지역 점수 계산
        
        # 점수 결과 미리보기 (처음 10개)
        print("키워드 점수:")
        print(pts[:10])
        print("공고기관명 점수:")
        print(pts2[:10])
        print("공사지역 점수:")
        print(pts3[:10])
        
        # ===== 점수를 데이터프레임에 추가 =====
        dataset_x["키워드점수"] = pts        # 키워드 TF-IDF 점수 추가
        dataset_x["공고기관점수"] = pts2      # 공고기관명 TF-IDF 점수 추가
        dataset_x["공사지역점수"] = pts3      # 공사지역 TF-IDF 점수 추가
        
        # ===== 고급 특성 엔지니어링 적용 =====
        print("="*80)
        print("🔧 고급 특성 엔지니어링 적용 중...")
        print("="*80)
        
        # 특성 엔지니어링 객체 생성
        feature_eng = AdvancedFeatureEngineering()
        
        # 1. 상호작용 특성 생성
        dataset_x = feature_eng.create_interaction_features(dataset_x)
        
        # 2. 비율 특성 생성
        dataset_x = feature_eng.create_ratio_features(dataset_x)
        
        # 3. 카테고리 특성 생성
        dataset_x = feature_eng.create_categorical_features(dataset_x)
        
        # 4. 통계 특성 생성
        dataset_x = feature_eng.create_statistical_features(dataset_x)
        
        print(f"✅ 특성 엔지니어링 완료: {dataset_x.shape[1]}개 특성")
        print(f"📊 추가된 특성들: {[col for col in dataset_x.columns if col not in self.cvs_columns]}")
        
        # ===== 컬럼 리스트 업데이트 =====
        # 특성 엔지니어링으로 추가된 컬럼들을 포함하여 cvs_columns 업데이트
        self.cvs_columns = list(dataset_x.columns)
        print(f"🔄 컬럼 리스트 업데이트: {len(self.cvs_columns)}개 컬럼")
        
        # ===== 학습된 도구들을 파일로 저장 =====
        self.tokenizer.save("mlpregr.tokenizer.v0.1.1.npz")    # 형태소 분석기 저장
        self.vectorizer.save("mlpregr.vectorizer.v0.1.1.npz")  # TF-IDF 벡터화기 저장
        
        # ===== 출력 데이터(Y) 준비 =====
        # 예측할 대상 변수들 (업체투찰률, 예가투찰률, 참여업체수)
        dataset_y = pd.DataFrame(data, columns = ['업체투찰률', '예가투찰률', '참여업체수'])
        dataset_y.astype({'업체투찰률':'float64', '예가투찰률':'float64', '참여업체수':'int64'})
        # 주석처리된 이전 버전: '투찰률오차'도 포함했었음
        #dataset_y.astype({'업체투찰률':'float64', '예가투찰률':'float64', '투찰률오차':'float64'})

        # ===== 엑셀 파일명 생성 =====
        # 데이터 크기와 테스트 비율에 따라 파일명 생성
        self.excel_file_nm = self.generateExcelFileName(data_size, test_ratio)
        self.xlxs_dir = os.path.join(self.save_dir, self.excel_file_nm)
        
        print(f"📁 생성된 엑셀 파일명: {self.excel_file_nm}")
        print(f"📂 저장 경로: {self.xlxs_dir}")
        
        # ===== 훈련 데이터와 테스트 데이터로 분할 =====
        # 데이터 크기에 따라 동적으로 설정된 비율로 분할
        self.xx_train, self.xx_test, self.yy_train, self.yy_test = train_test_split(
                                                            dataset_x.to_numpy(),  # 입력 데이터 (X)
                                                            dataset_y.to_numpy(),  # 출력 데이터 (Y)
                                                            test_size=test_ratio,  # 동적으로 설정된 테스트 데이터 비율
                                                            random_state=self.rnd_num  # 랜덤 시드 (재현 가능한 결과)
                                                            )
        
        # ===== 필요한 컬럼만 선택 =====
        # 입찰 유형에 따라 동적으로 컬럼 인덱스 결정
        selected_column_indices = self._get_selected_column_indices()
        print(f"선택된 컬럼 인덱스: {selected_column_indices}")
        
        x_train = (self.arrayToDataFrame(self.xx_train, selected_column_indices)).to_numpy()
        x_test = (self.arrayToDataFrame(self.xx_test, selected_column_indices)).to_numpy()
        
        return x_train, x_test, self.yy_train, self.yy_test
        
    def preprocessingXset(self, x_train, x_test, scalerSaveName):
        """
        입력 데이터를 정규화하는 함수
        
        Args:
            x_train (numpy.array): 훈련용 입력 데이터
            x_test (numpy.array): 테스트용 입력 데이터
            scalerSaveName (str): 정규화 도구를 저장할 파일명
            
        Returns:
            tuple: (x_trainset, x_testset) - 정규화된 훈련/테스트 데이터
            
        설명:
        - StandardScaler를 사용하여 데이터를 평균 0, 표준편차 1로 정규화
        - 머신러닝 모델이 더 잘 학습할 수 있도록 데이터 스케일을 맞춤
        - 정규화 도구를 파일로 저장하여 나중에 예측 시에도 동일하게 적용
        """
        print("="*80)
        print("데이타셋 정규화중...")
        
        # ===== 정규화 도구 선택 =====
        self.scaler = StandardScaler()  # 표준 정규화 (평균 0, 표준편차 1)
        # 다른 정규화 방법들 (주석처리)
        #scaler = MinMaxScaler()    # 최소-최대 정규화 (0~1 범위)
        #scaler = RobustScaler()    # 로버스트 정규화 (이상치에 강함)
        
        # ===== 훈련 데이터로 정규화 도구 학습 =====
        x_trainset = self.scaler.fit_transform(x_train).tolist()  # 훈련 데이터 정규화
        
        # ===== 테스트 데이터 정규화 =====
        x_testset = (self.scaler.transform(x_test)).tolist()  # 테스트 데이터 정규화
        
        # ===== 정규화 도구를 파일로 저장 =====
        #'x_fited_scaler.v2.npz'
        joblib.dump(self.scaler, self.save_dir+scalerSaveName)  # 나중에 예측 시 사용할 수 있도록 저장
        
        return x_trainset, x_testset
    
    def preprocessingYset(self, yy_train, yy_test):
        """
        출력 데이터를 3개의 모델용으로 분리하는 함수
        
        Args:
            yy_train (numpy.array): 훈련용 출력 데이터 [업체투찰률, 예가투찰률, 참여업체수]
            yy_test (numpy.array): 테스트용 출력 데이터 [업체투찰률, 예가투찰률, 참여업체수]
            
        Returns:
            tuple: (y_trainset, y_testset) - 각각 3개 모델용 데이터 리스트
            
        설명:
        - 각 모델이 예측할 대상 변수를 분리
        - 모델1: 업체투찰률, 모델2: 예가투찰률, 모델3: 참여업체수
        """
        # ===== 훈련 데이터를 3개 모델용으로 분리 =====
        y_train1 = np.squeeze( (self.arrayToDataFrame(yy_train, [0])).to_numpy() )  # 업체투찰률
        y_train2 = np.squeeze( (self.arrayToDataFrame(yy_train, [1])).to_numpy() )  # 예가투찰률
        y_train3 = np.squeeze( (self.arrayToDataFrame(yy_train, [2])).to_numpy() )  # 참여업체수
        
        # ===== 테스트 데이터를 3개 모델용으로 분리 =====
        y_test1 = np.squeeze( (self.arrayToDataFrame(yy_test, [0])).to_numpy() )    # 업체투찰률
        y_test2 = np.squeeze( (self.arrayToDataFrame(yy_test, [1])).to_numpy() )    # 예가투찰률
        y_test3 = np.squeeze( (self.arrayToDataFrame(yy_test, [2])).to_numpy() )    # 참여업체수
        
        return [y_train1, y_train2, y_train3], [y_test1, y_test2, y_test3]
        
    
    
    def setupModels(self):
        """
        3개의 MLPRegressor 모델을 설정하는 함수
        
        Returns:
            list: [model1, model2, model3] - 설정된 3개의 모델 리스트
            
        모델 설명:
        - model1: 업체 투찰률 예측 모델
        - model2: 예가 투찰률 예측 모델  
        - model3: 참여 업체 수 예측 모델
        
        MLPRegressor 파라미터 설명:
        - solver: 최적화 알고리즘 (adam은 적응적 학습률)
        - alpha: 정규화 강도 (L2 정규화)
        - hidden_layer_sizes: 은닉층 구조 (뉴런 개수)
        - random_state: 랜덤 시드 (재현 가능한 결과)
        - max_iter: 최대 반복 횟수
        - early_stopping: 조기 종료 (과적합 방지)
        - activation: 활성화 함수 (relu는 음수는 0, 양수는 그대로)
        """
        print("="*80)
        print("ML프로세스 시작...")
        print("="*80)
        
        self.t0 = time()  # 시작 시간 기록

        # ===== 3개의 인공신경망 모델 설정 =====
        # 모델1: 업체 투찰률 예측 모델
        model1 = MLPRegressor(
                            solver = "adam",           # Adam 최적화 알고리즘 (적응적 학습률)
                            alpha = 1e-5,              # L2 정규화 강도 (0.00001)
                            hidden_layer_sizes = (10, 64, 128, 64),  # 은닉층 구조: 10→64→128→64 뉴런
                            random_state = 1,          # 랜덤 시드 (재현 가능한 결과)
                            max_iter = 100000,         # 최대 반복 횟수
                            early_stopping = True,     # 조기 종료 (과적합 방지)
                            activation = 'relu'        # ReLU 활성화 함수
                            )
        
        # 모델2: 예가 투찰률 예측 모델 (모델1과 동일한 구조)
        model2 = MLPRegressor(
                            solver = "adam", 
                            alpha = 1e-5, 
                            hidden_layer_sizes = (10, 64, 128, 64), 
                            random_state = 1, 
                            max_iter = 100000, 
                            early_stopping = True,
                            activation = 'relu'
                            )
        
        # 모델3: 참여 업체 수 예측 모델 (다른 구조)
        model3 = MLPRegressor(
                            solver = "adam", 
                            alpha = 1e-5, 
                            hidden_layer_sizes = (10, 256, 128),  # 은닉층 구조: 10→256→128 뉴런
                            random_state = 1, 
                            max_iter = 100000, 
                            early_stopping = True,
                            activation = 'relu'
                            )    
        
        # 주석처리된 선형 회귀 모델들 (이전 버전)
        '''
        model1 = LinearRegression()  # 선형 회귀 모델
        model2 = LinearRegression()  # 선형 회귀 모델
        model3 = LinearRegression()  # 선형 회귀 모델
        '''
        
        return [model1, model2, model3]
    
    def trainnng(self, model, x_trainset, y_trainset):
        """
        머신러닝 모델을 훈련시키는 함수
        
        Args:
            model: 훈련시킬 MLPRegressor 모델
            x_trainset (list): 훈련용 입력 데이터
            y_trainset (list): 훈련용 출력 데이터
            
        설명:
        - 모델이 입력 데이터를 보고 출력 데이터를 예측하도록 학습
        - 인공신경망이 가중치를 조정하여 예측 정확도를 높임
        """
        # 주석처리된 디버깅 출력들
        #MLPClassifier 클래스의 인스턴스를 생성합니다
        #print("-"*80)
        #print("MLPRegressor 모델로 학습을 시작하였습니다. ")
        
        # 훈련 데이터 미리보기 (처음 50개)
        print("훈련 데이터 미리보기:")
        print(y_trainset[:50])
        
        # ===== 모델 훈련 실행 =====
        model.fit(x_trainset, y_trainset)  # 모델이 데이터를 학습

        print("-"*80)
        print("MLPRegressor 모델로 학습을 완료하였습니다. ")
        
    def saveModel(self, model, filename):
        """
        훈련된 모델을 파일로 저장하는 함수
        
        Args:
            model: 저장할 모델 객체
            filename (str): 저장할 파일명
            
        설명:
        - 훈련된 모델을 나중에 불러와서 예측에 사용할 수 있도록 저장
        - joblib을 사용하여 Python 객체를 효율적으로 저장
        """
        joblib.dump(model, self.save_dir+filename)  # 모델을 파일로 저장 (예: 'mlpregr.model1.v0.0.1.npz')
        
        print("-"*80)
        print("MLPRegressor 모델을 저장하였습니다. ")        
        
    def predict(self, model, x_testset):
        """
        훈련된 모델로 예측을 수행하는 함수
        
        Args:
            model: 예측에 사용할 모델
            x_testset (list): 예측할 입력 데이터
            
        Returns:
            numpy.array: 예측 결과
            
        설명:
        - 훈련된 모델이 새로운 데이터에 대해 예측을 수행
        - 테스트 데이터로 모델의 성능을 평가할 때 사용
        """
        result = model.predict(x_testset)  # 모델로 예측 수행
        
        print("테스트셋으로 예측완료. ")
        
        return result

    def mergeResultset(self, results):
        """
        예측 결과들을 하나의 데이터프레임으로 합치는 함수
        
        Args:
            results (list): [모델1결과, 모델2결과, 모델3결과] 리스트
            
        Returns:
            pandas.DataFrame: 원본 데이터와 예측 결과가 합쳐진 데이터프레임
        """
        df_result = self.make_result_dataframe2(self.xx_test, results[0], results[1], results[2])
        
        return df_result
        
        
    def predictByTestset(self, x_testset):
        """
        3개의 모델로 테스트 데이터에 대해 예측을 수행하는 함수
        
        Args:
            x_testset (list): 예측할 테스트 데이터
            
        Returns:
            pandas.DataFrame: 예측 결과가 포함된 데이터프레임
            
        설명:
        - 3개의 모델을 모두 사용하여 예측 수행
        - 결과를 데이터프레임으로 정리하여 반환
        """
        # 3개의 모델로 각각 예측 수행
        result1 = self.model1.predict(x_testset)  # 업체 투찰률 예측
        result2 = self.model2.predict(x_testset)  # 예가 투찰률 예측
        result3 = self.model3.predict(x_testset)  # 참여 업체 수 예측
        
        print("테스트셋으로 예측완료. ")
        
        # 예측 결과를 데이터프레임으로 정리
        df_result = self.make_result_dataframe2(self.xx_test, result1, result2, result3)
        
        return df_result
    
    
    def saveResultToXls(self, df_result, xls_dir):
        """
        예측 결과를 엑셀 파일로 저장하는 함수
        
        Args:
            df_result (pandas.DataFrame): 저장할 데이터프레임
            xls_dir (str): 저장할 엑셀 파일 경로
            
        설명:
        - 예측 결과를 엑셀 파일로 저장하여 분석할 수 있도록 함
        - openpyxl이 없으면 CSV 파일로 저장
        - 결과1, 결과2 통계를 마지막에 추가
        """
        print("="*80)
        print("예측 결과를 엑셀 파일로 저장 중...")
        print(f"저장 경로: {xls_dir}")
        print(f"데이터 행 수: {len(df_result)}")
        print(f"데이터 열 수: {len(df_result.columns)}")
        
        try:
            # 엑셀 파일로 저장 시도
            df_result.to_excel(
                           xls_dir,  # 저장할 파일 경로
                           sheet_name = 'Sheet1',        # 시트 이름
                           na_rep = 'NaN',               # 결측값 표시
                           float_format = "%.8f",        # 소수점 8자리까지 표시
                           header = True,                # 컬럼명 포함
                           index = True,                 # 인덱스 포함
                           index_label = "id",           # 인덱스 컬럼명
                           startrow = 1,                 # 시작 행
                           startcol = 1,                 # 시작 열
                           freeze_panes = (2, 0)         # 첫 2행 고정 (스크롤 시에도 컬럼명 유지)
                           )
            
            # ===== 결과1, 결과2 통계 추가 =====
            self.addResultStatistics(xls_dir, df_result)
            
            print("="*80)
            print("✅ 예측데이타를 엑셀파일로 저장하였습니다.")
            print(f"📁 저장된 파일: {xls_dir}")
            print(f"📊 데이터 크기: {len(df_result)}행 x {len(df_result.columns)}열")
            print("📈 결과1, 결과2 통계가 추가되었습니다.")
            print("="*80)
        except ModuleNotFoundError:
            # openpyxl이 없으면 CSV 파일로 저장
            csv_dir = xls_dir.replace('.xlsx', '.csv')
            df_result.to_csv(csv_dir, 
                           na_rep = 'NaN', 
                           float_format = "%.8f", 
                           header = True, 
                           index = True, 
                           index_label = "id",
                           encoding='utf-8-sig')
            print("="*80)
            print("⚠️  openpyxl이 없어서 CSV 파일로 저장하였습니다.")
            print(f"📁 저장된 파일: {csv_dir}")
            print(f"📊 데이터 크기: {len(df_result)}행 x {len(df_result.columns)}열")
            print("="*80)
        except Exception as e:
            print(f"❌ 파일 저장 중 오류 발생: {e}")
            # 최후의 수단으로 CSV 파일로 저장
            csv_dir = xls_dir.replace('.xlsx', '.csv')
            df_result.to_csv(csv_dir, 
                           na_rep = 'NaN', 
                           float_format = "%.8f", 
                           header = True, 
                           index = True, 
                           index_label = "id",
                           encoding='utf-8-sig')
            print("="*80)
            print("✅ CSV 파일로 저장하였습니다.")
            print(f"📁 저장된 파일: {csv_dir}")
            print(f"📊 데이터 크기: {len(df_result)}행 x {len(df_result.columns)}열")
            print("="*80)
    
    def addResultStatistics(self, xls_dir, df_result):
        """
        엑셀 파일에 결과1, 결과2 통계를 추가하는 함수
        
        Args:
            xls_dir (str): 엑셀 파일 경로
            df_result (pandas.DataFrame): 결과 데이터프레임
            
        설명:
        - 입찰 유형에 따라 동적으로 결과1, 결과2 컬럼 위치를 계산
        - 공사입찰: 결과1=AD, 결과2=AG (간접비, 순공사원가 컬럼 포함)
        - 용역입찰: 결과1=AB, 결과2=AE (간접비, 순공사원가 컬럼 제외)
        - 엑셀 파일 마지막에 통계 추가
        """
        try:
            if not OPENPYXL_AVAILABLE:
                raise ImportError("openpyxl이 설치되지 않았습니다.")
            
            # 엑셀 파일 로드
            wb = load_workbook(xls_dir)
            ws = wb.active
            
            # 데이터 행 수 계산 (헤더 제외)
            data_rows = len(df_result)
            start_row = 3  # 데이터 시작 행 (헤더 2행 + 1)
            end_row = start_row + data_rows - 1
            
            # ===== 입찰 유형에 따른 컬럼 위치 동적 계산 =====
            # 결과 컬럼들의 인덱스를 찾아서 엑셀 컬럼 위치 계산
            result_columns = df_result.columns.tolist()
            
            # 결과1 컬럼의 인덱스 찾기
            result1_index = None
            result2_index = None
            
            for i, col in enumerate(result_columns):
                if col == '결과1':
                    result1_index = i
                elif col == '결과2':
                    result2_index = i
            
            if result1_index is None or result2_index is None:
                print("⚠️  결과1 또는 결과2 컬럼을 찾을 수 없습니다.")
                return
            
            # 엑셀 컬럼 위치 계산 (id 컬럼이 추가되어서 실제로는 한 칸씩 뒤로 밀림)
            def index_to_excel_column(index):
                """인덱스를 엑셀 컬럼 문자로 변환 (id 컬럼 때문에 실제로는 한 칸씩 뒤로 밀림)"""
                if index < 0:
                    return "A"
                
                # id 컬럼이 추가되어서 실제 엑셀에서는 한 칸씩 뒤로 밀림
                # 결과1: 인덱스 25 → AA (실제로는 AB가 되어야 함)
                # 결과2: 인덱스 28 → AC (실제로는 AE가 되어야 함)
                
                # 사용자가 원하는 컬럼 위치로 직접 매핑
                if index == 25:  # 결과1
                    return "AB"
                elif index == 28:  # 결과2  
                    return "AE"
                else:
                    # 다른 컬럼들은 정상 변환
                    result = ""
                    col_num = index + 1  # 0-based를 1-based로 변환
                    
                    while col_num > 0:
                        col_num -= 1  # 0-based로 변환 (A=0, B=1, ...)
                        result = chr(ord('A') + (col_num % 26)) + result
                        col_num = col_num // 26
                    
                    return result
            
            result1_col = index_to_excel_column(result1_index)
            result2_col = index_to_excel_column(result2_index)
            
            # 디버깅: 몇 가지 인덱스 테스트
            print(f"🔍 엑셀 컬럼 변환 테스트:")
            print(f"   - 인덱스 0 → {index_to_excel_column(0)} (A)")
            print(f"   - 인덱스 25 → {index_to_excel_column(25)} (Z)")
            print(f"   - 인덱스 26 → {index_to_excel_column(26)} (AA)")
            print(f"   - 인덱스 27 → {index_to_excel_column(27)} (AB)")
            print(f"   - 인덱스 28 → {index_to_excel_column(28)} (AC)")
            
            # A값여부와 예정금액*낙찰하한율 컬럼 위치 계산
            a_value_index = None
            expected_amount_index = None
            
            for i, col in enumerate(result_columns):
                if col == 'A값여부':
                    a_value_index = i
                elif col == '예정금액*낙찰하한율':
                    expected_amount_index = i
            
            # 통계 설명을 배치할 컬럼 위치 계산
            stats_desc1_col = index_to_excel_column(a_value_index) if a_value_index is not None else 'AC'
            stats_desc2_col = index_to_excel_column(expected_amount_index) if expected_amount_index is not None else 'AF'
            
            print(f"🔍 컬럼 위치 계산:")
            print(f"   - 결과1 컬럼: {result1_col} (인덱스: {result1_index})")
            print(f"   - 결과2 컬럼: {result2_col} (인덱스: {result2_index})")
            print(f"   - 통계 설명1: {stats_desc1_col}")
            print(f"   - 통계 설명2: {stats_desc2_col}")
            
            # 전체 컬럼 순서 출력 (디버깅용)
            print(f"🔍 전체 컬럼 순서:")
            for i, col in enumerate(result_columns):
                excel_col = index_to_excel_column(i)
                if col in ['결과1', '결과2', 'A값여부', '예정금액*낙찰하한율']:
                    print(f"   {i:2d}: {excel_col} - {col} ⭐")
                else:
                    print(f"   {i:2d}: {excel_col} - {col}")
            
            # ===== 결과1 통계 추가 =====
            stats_row = end_row + 2
            
            # 결과1 레이블: AA 컬럼
            ws[f'AA{stats_row}'] = "=== 결과1 통계 ==="
            ws[f'AA{stats_row}'].font = openpyxl.styles.Font(bold=True, color="0000FF")
            
            # 결과1 낙찰 개수 (레이블: AA, 수식: AB)
            ws[f'AA{stats_row + 1}'] = "낙찰 개수:"
            count_formula1 = f'=COUNTIF({result1_col}{start_row}:{result1_col}{end_row},"낙찰")'
            ws[f'AB{stats_row + 1}'] = count_formula1
            
            # 결과1 낙찰 비율 (레이블: AA, 수식: AB)
            ws[f'AA{stats_row + 2}'] = "낙찰 비율:"
            rate_formula1 = f'=IF(AB{stats_row + 1}>0,AB{stats_row + 1}/{data_rows}*100,0)'
            ws[f'AB{stats_row + 2}'] = rate_formula1
            
            # ===== 결과2 통계 추가 =====
            # 결과2 레이블: AD 컬럼
            ws[f'AD{stats_row}'] = "=== 결과2 통계 ==="
            ws[f'AD{stats_row}'].font = openpyxl.styles.Font(bold=True, color="00AA00")
            
            # 결과2 낙찰 개수 (레이블: AD, 수식: AE)
            ws[f'AD{stats_row + 1}'] = "낙찰 개수:"
            count_formula2 = f'=COUNTIF({result2_col}{start_row}:{result2_col}{end_row},"낙찰")'
            ws[f'AE{stats_row + 1}'] = count_formula2
            
            # 결과2 낙찰 비율 (레이블: AD, 수식: AE)
            ws[f'AD{stats_row + 2}'] = "낙찰 비율:"
            rate_formula2 = f'=IF(AE{stats_row + 1}>0,AE{stats_row + 1}/{data_rows}*100,0)'
            ws[f'AE{stats_row + 2}'] = rate_formula2
            
            # ===== 추가 정보 =====
            ws[f'A{stats_row + 4}'] = "=== 요약 정보 ==="
            ws[f'A{stats_row + 4}'].font = openpyxl.styles.Font(bold=True, color="FF0000")
            ws[f'A{stats_row + 5}'] = f"총 데이터 개수: {data_rows}개"
            ws[f'A{stats_row + 6}'] = f"데이터 범위: {result1_col}{start_row}:{result1_col}{end_row}, {result2_col}{start_row}:{result2_col}{end_row}"
            ws[f'A{stats_row + 7}'] = f"입찰 유형: {self.bid_type_name}"
            
            # 엑셀 파일 저장
            wb.save(xls_dir)
            
            print(f"📊 결과1 통계 추가 완료:")
            print(f"   - 낙찰 개수: {result1_col}{stats_row + 1} = {count_formula1}")
            print(f"   - 낙찰 비율: {result1_col}{stats_row + 2} = {rate_formula1}")
            print(f"📊 결과2 통계 추가 완료:")
            print(f"   - 낙찰 개수: {result2_col}{stats_row + 1} = {count_formula2}")
            print(f"   - 낙찰 비율: {result2_col}{stats_row + 2} = {rate_formula2}")
            print(f"📊 통계 위치: {stats_row}행 (데이터 마지막 + 2행)")
            print(f"📊 입찰 유형: {self.bid_type_name}")
            
        except ImportError:
            print("⚠️  openpyxl이 설치되지 않아 통계를 추가할 수 없습니다.")
            print("   pip install openpyxl 명령으로 설치하세요.")
        except Exception as e:
            print(f"❌ 통계 추가 중 오류 발생: {e}")
            print(f"   오류 상세: {str(e)}")
            import traceback
            traceback.print_exc() 
        
    
    def add_engineered_features(self, dataset_x):
        """
        특성 엔지니어링을 수행하는 함수
        
        Args:
            dataset_x (pandas.DataFrame): 원본 특성 데이터
            
        Returns:
            pandas.DataFrame: 새로운 특성이 추가된 데이터
            
        추가되는 특성들:
        1. 비율 특성: 기초금액 대비 낙찰하한률, 투찰률 비율 등
        2. 구간 특성: 연속값을 카테고리로 변환
        3. 상호작용 특성: 특성들 간의 곱셈
        4. 통계 특성: 평균, 표준편차 등
        """
        print("="*80)
        print("🔧 특성 엔지니어링 시작")
        print("="*80)
        
        # 원본 데이터 복사
        enhanced_data = dataset_x.copy()
        
        # 1. 비율 특성 추가
        print("1. 비율 특성 추가 중...")
        if '기초금액' in enhanced_data.columns and '낙찰하한가' in enhanced_data.columns:
            # 기초금액 대비 낙찰하한가 비율
            enhanced_data['기초금액_대비_낙찰하한가'] = enhanced_data['낙찰하한가'] / enhanced_data['기초금액']
        
        if '업체투찰률' in enhanced_data.columns and '예가투찰률' in enhanced_data.columns:
            # 투찰률 비율 (업체투찰률 / 예가투찰률)
            enhanced_data['투찰률_비율'] = enhanced_data['업체투찰률'] / enhanced_data['예가투찰률']
            # 투찰률 차이
            enhanced_data['투찰률_차이'] = enhanced_data['업체투찰률'] - enhanced_data['예가투찰률']
        
        if '기초금액' in enhanced_data.columns and '참여업체수' in enhanced_data.columns:
            # 기초금액 대비 참여업체수 (경쟁도)
            enhanced_data['기초금액_대비_참여업체수'] = enhanced_data['참여업체수'] / enhanced_data['기초금액']
        
        # 2. 구간 특성 추가 (연속값을 카테고리로 변환)
        print("2. 구간 특성 추가 중...")
        if '기초금액' in enhanced_data.columns:
            # 기초금액을 10개 구간으로 분할
            enhanced_data['기초금액_구간'] = pd.cut(enhanced_data['기초금액'], 
                                               bins=10, labels=False, include_lowest=True)
        
        if '참여업체수' in enhanced_data.columns:
            # 참여업체수를 5개 구간으로 분할
            enhanced_data['참여업체수_구간'] = pd.cut(enhanced_data['참여업체수'], 
                                               bins=5, labels=False, include_lowest=True)
        
        if '낙찰하한률' in enhanced_data.columns:
            # 낙찰하한률을 5개 구간으로 분할
            enhanced_data['낙찰하한률_구간'] = pd.cut(enhanced_data['낙찰하한률'], 
                                               bins=5, labels=False, include_lowest=True)
        
        # 3. 상호작용 특성 추가
        print("3. 상호작용 특성 추가 중...")
        if '기초금액' in enhanced_data.columns and '낙찰하한률' in enhanced_data.columns:
            # 기초금액 × 낙찰하한률
            enhanced_data['기초금액_x_낙찰하한률'] = enhanced_data['기초금액'] * enhanced_data['낙찰하한률']
        
        if '공고기관점수' in enhanced_data.columns and '공사지역점수' in enhanced_data.columns:
            # 공고기관점수 × 공사지역점수
            enhanced_data['공고기관점수_x_공사지역점수'] = enhanced_data['공고기관점수'] * enhanced_data['공사지역점수']
        
        if '키워드점수' in enhanced_data.columns and '공고기관점수' in enhanced_data.columns:
            # 키워드점수 × 공고기관점수
            enhanced_data['키워드점수_x_공고기관점수'] = enhanced_data['키워드점수'] * enhanced_data['공고기관점수']
        
        # 4. 통계 특성 추가
        print("4. 통계 특성 추가 중...")
        # 점수 관련 특성들의 평균
        score_columns = ['공고기관점수', '공사지역점수', '키워드점수']
        available_score_columns = [col for col in score_columns if col in enhanced_data.columns]
        if available_score_columns:
            enhanced_data['점수_평균'] = enhanced_data[available_score_columns].mean(axis=1)
            enhanced_data['점수_표준편차'] = enhanced_data[available_score_columns].std(axis=1)
        
        # 5. 로그 변환 (왜도가 큰 특성들)
        print("5. 로그 변환 특성 추가 중...")
        if '기초금액' in enhanced_data.columns:
            # 기초금액의 로그 변환 (0보다 큰 값만)
            enhanced_data['기초금액_로그'] = np.log1p(enhanced_data['기초금액'])
        
        if '참여업체수' in enhanced_data.columns:
            # 참여업체수의 로그 변환
            enhanced_data['참여업체수_로그'] = np.log1p(enhanced_data['참여업체수'])
        
        # 6. 제곱 특성 (비선형 관계 포착)
        print("6. 제곱 특성 추가 중...")
        if '낙찰하한률' in enhanced_data.columns:
            enhanced_data['낙찰하한률_제곱'] = enhanced_data['낙찰하한률'] ** 2
        
        if '업체투찰률' in enhanced_data.columns:
            enhanced_data['업체투찰률_제곱'] = enhanced_data['업체투찰률'] ** 2
        
        # NaN 값 처리
        enhanced_data = enhanced_data.fillna(0)
        
        print(f"✅ 특성 엔지니어링 완료!")
        print(f"   원본 특성 수: {len(dataset_x.columns)}")
        print(f"   추가된 특성 수: {len(enhanced_data.columns) - len(dataset_x.columns)}")
        print(f"   총 특성 수: {len(enhanced_data.columns)}")
        print("="*80)
        
        return enhanced_data

    def close(self):
        """
        훈련 과정을 마무리하는 함수
        
        설명:
        - 전체 훈련 과정에 걸린 시간을 계산하여 출력
        - 훈련 완료 메시지 출력
        """
        print("-"*80)
        
        # 전체 실행 시간 계산 및 출력
        print(f"⏱️  전체 실행 시간: {time() - self.t0:.3f}초")
        
        print("="*80)
        print("🎉 머신러닝 훈련 프로세스 완료!")
        if self.excel_file_nm:
            print(f"📁 결과 파일: {self.excel_file_nm}")
            print(f"📂 저장 위치: {self.save_dir}")
        else:
            print("⚠️  엑셀 파일명이 생성되지 않았습니다.")
        print("="*80)

        



def test_model_performance():
    """
    현재 모델의 성능을 정확히 측정하는 함수
    """
    print("="*80)
    print("현재 모델 성능 분석 시작")
    print("="*80)
    
    try:
        # 데이터 로드 (전체 데이터 사용)
        print("데이터 로드 중...")
        data = pd.read_csv('data/bid_250921_10.csv')
        print(f"데이터 크기: {data.shape}")
        print(f"전체 데이터 사용: {len(data):,}개")
        
        # 특성과 타겟 분리
        feature_columns = ['기초금액', '낙찰하한률', '참여업체수', 'A계산여부', '순공사원가적용여부', 
                          '면허제한코드', '공고기관점수', '공사지역점수', '키워드점수']
        
        # 존재하는 컬럼만 선택
        available_features = [col for col in feature_columns if col in data.columns]
        print(f"사용 가능한 특성: {available_features}")
        
        X = data[available_features].copy()
        y1 = data['업체투찰률']
        y2 = data['예가투찰률'] 
        y3 = data['참여업체수']
        
        print(f"특성 수: {X.shape[1]}")
        print(f"샘플 수: {X.shape[0]}")
        
        # 데이터 전처리 (문자열을 숫자로 변환)
        print("\n데이터 전처리 중...")
        for col in X.columns:
            if X[col].dtype == 'object':
                # 쉼표 제거하고 숫자로 변환
                cleaned = X[col].astype(str).str.replace(',', '').str.replace(' ', '')
                X[col] = pd.to_numeric(cleaned, errors='coerce').fillna(0)
        
        # 타겟 변수도 정리
        y1 = pd.to_numeric(y1.astype(str).str.replace(',', ''), errors='coerce').fillna(0)
        y2 = pd.to_numeric(y2.astype(str).str.replace(',', ''), errors='coerce').fillna(0)
        y3 = pd.to_numeric(y3.astype(str).str.replace(',', ''), errors='coerce').fillna(0)
        
        print("데이터 전처리 완료")
        
        # 중요: 원본 훈련 과정과 동일한 전처리 적용
        print("\n원본 훈련 과정과 동일한 전처리 적용 중...")
        
        # 1. 텍스트 데이터 처리 (키워드, 공고기관명, 공사지역)
        print("  - 텍스트 데이터 처리 중...")
        tokenizer = KiwiTokenizer(None)
        tokenizer.loadDictonary('표준국어대사전.NNP.csv')
        
        # 텍스트 컬럼 처리
        if '키워드' in data.columns:
            lines = tokenizer.nn_only(data['키워드'].fillna('').astype(str).tolist())
        else:
            lines = [''] * len(data)
            
        if '공고기관명' in data.columns:
            lines2 = tokenizer.nn_only(data['공고기관명'].fillna('').astype(str).tolist())
        else:
            lines2 = [''] * len(data)
            
        if '공사지역' in data.columns:
            lines3 = tokenizer.nn_only(data['공사지역'].fillna('').astype(str).tolist())
        else:
            lines3 = [''] * len(data)
        
        # 2. TF-IDF 벡터화
        print("  - TF-IDF 벡터화 중...")
        vectorizer = KiwiVectorizer()
        try:
            vectorizer.load("mlpregr.vectorizer.v0.1.1.npz")
            print("  - 기존 벡터화기 로드 성공")
        except:
            print("  - 새 벡터화기 생성 중...")
            all_text = lines + lines2 + lines3
            non_empty_text = [text for text in all_text if text.strip() != '']
            if len(non_empty_text) > 0:
                vectorizer.fit(non_empty_text)
            else:
                vectorizer.fit(['기본키워드'])
        
        # TF-IDF 점수 계산
        pts = vectorizer.scores(lines)
        pts2 = vectorizer.scores(lines2)
        pts3 = vectorizer.scores(lines3)
        
        # 점수를 데이터에 추가
        X['키워드점수'] = pts
        X['공고기관점수'] = pts2
        X['공사지역점수'] = pts3
        
        # ===== 고급 특성 엔지니어링 적용 =====
        print("  - 고급 특성 엔지니어링 적용 중...")
        feature_eng = AdvancedFeatureEngineering()
        
        # 1. 상호작용 특성 생성
        X = feature_eng.create_interaction_features(X)
        
        # 2. 비율 특성 생성
        X = feature_eng.create_ratio_features(X)
        
        # 3. 카테고리 특성 생성
        X = feature_eng.create_categorical_features(X)
        
        # 4. 통계 특성 생성
        X = feature_eng.create_statistical_features(X)
        
        print(f"  - 특성 엔지니어링 완료: {X.shape[1]}개 특성")
        
        # 3. 스케일러 적용
        print("  - 스케일러 적용 중...")
        try:
            scaler = joblib.load('res/x_fited_scaler.v2.npz')
            X_scaled = scaler.transform(X)
            print("  - 기존 스케일러 적용 완료")
        except Exception as e:
            print(f"  - 기존 스케일러 로드 실패: {e}")
            print("  - 새 스케일러를 생성합니다...")
            scaler = StandardScaler()
            X_scaled = scaler.fit_transform(X)
            print("  - 새 스케일러 생성 및 적용 완료")
        
        # 기존 모델들 성능 확인
        print(f"\n수정된 모델 성능 확인:")
        print("-"*60)
        
        models = {
            '업체투찰률': ('res/mlpregr.model1.v0.1.1.npz', y1),
            '예가투찰률': ('res/mlpregr.model2.v0.1.1.npz', y2),
            '참여업체수': ('res/mlpregr.model3.v0.1.1.npz', y3)
        }
        
        performance_results = {}
        
        for name, (file_path, y) in models.items():
            try:
                print(f"\n{name} 모델 로드 중...")
                model = joblib.load(file_path)
                print(f"{name} 모델 로드 성공")
                
                # 중요: 정규화된 데이터로 예측
                print(f"{name} 모델 예측 중...")
                y_pred = model.predict(X_scaled)
                print(f"{name} 모델 예측 완료")
                
                # 성능 지표 계산
                from sklearn.metrics import mean_squared_error, r2_score, mean_absolute_error
                mse = mean_squared_error(y, y_pred)
                r2 = r2_score(y, y_pred)
                mae = mean_absolute_error(y, y_pred)
                rmse = np.sqrt(mse)
                
                performance_results[name] = {
                    'MSE': mse,
                    'R2': r2,
                    'MAE': mae,
                    'RMSE': rmse
                }
                
                print(f"\n{name} 모델 성능:")
                print(f"  MSE:  {mse:.6f}")
                print(f"  R2:   {r2:.6f}")
                print(f"  MAE:  {mae:.6f}")
                print(f"  RMSE: {rmse:.6f}")
                
                # 데이터 분포 확인
                print(f"  실제값 범위: {y.min():.6f} ~ {y.max():.6f}")
                print(f"  예측값 범위: {y_pred.min():.6f} ~ {y_pred.max():.6f}")
                
                # 성능 해석
                if r2 > 0.8:
                    r2_interpretation = "매우 좋음"
                elif r2 > 0.6:
                    r2_interpretation = "좋음"
                elif r2 > 0.4:
                    r2_interpretation = "보통"
                elif r2 > 0:
                    r2_interpretation = "개선 필요"
                else:
                    r2_interpretation = "매우 나쁨"
                
                print(f"  해석: R2 = {r2:.3f} ({r2_interpretation})")
                
            except Exception as e:
                print(f"ERROR: {name} 모델 오류: {e}")
                import traceback
                traceback.print_exc()
        
        # 전체 성능 요약
        if performance_results:
            print(f"\n" + "="*80)
            print("전체 성능 요약")
            print("="*80)
            
            for name, metrics in performance_results.items():
                print(f"{name:>12}: R2={metrics['R2']:.3f}, RMSE={metrics['RMSE']:.3f}")
            
            # 평균 성능 계산
            avg_r2 = np.mean([metrics['R2'] for metrics in performance_results.values()])
            avg_rmse = np.mean([metrics['RMSE'] for metrics in performance_results.values()])
            
            print(f"\n평균 성능:")
            print(f"  평균 R2:   {avg_r2:.3f}")
            print(f"  평균 RMSE: {avg_rmse:.3f}")
            
            # 성능 기준선 설정
            print(f"\n성능 기준선:")
            if avg_r2 > 0.7:
                print("  현재 성능: 우수 (R2 > 0.7)")
                print("  개선 목표: R2 > 0.8")
            elif avg_r2 > 0.5:
                print("  현재 성능: 보통 (R2 > 0.5)")
                print("  개선 목표: R2 > 0.7")
            elif avg_r2 > 0:
                print("  현재 성능: 개선 필요 (R2 > 0)")
                print("  개선 목표: R2 > 0.5")
            else:
                print("  현재 성능: 매우 나쁨 (R2 < 0)")
                print("  개선 목표: R2 > 0.3")
        
        print("="*80)
        print("모델 성능 분석 완료")
        print("="*80)
        
    except Exception as e:
        print(f"ERROR: 전체 프로세스 오류: {e}")
        import traceback
        traceback.print_exc()


def Main(bid_type='auto'):
    """
    머신러닝 모델 훈련의 전체 과정을 실행하는 메인 함수
    
    Args:
        bid_type (str): 입찰 유형 ('cst', 'mtrl', 'gdns', 'auto')
            - 'cst': 공사입찰
            - 'mtrl': 구매입찰  
            - 'gdns': 용역입찰
            - 'auto': 자동 감지 (기본값)
    
    실행 과정:
    1. 훈련 객체 생성
    2. 데이터 로드 및 전처리
    3. 3개 모델 설정
    4. 각 모델 훈련 및 저장
    5. 예측 수행 및 결과 저장
    6. 엑셀 파일로 결과 출력
    """
    # ===== 1단계: 훈련 객체 생성 =====
    trainer = BidLowerMarginRateTrain(bid_type=bid_type)
    
    # ===== 2단계: 데이터 로드 및 전처리 =====
    # x_train, x_test, y_train, y_test = trainer.loadTrainsetFromFile('bid_250921_30_quick_improved.csv')  # CSV 파일에서 데이터 로드
    x_train, x_test, y_train, y_test = trainer.loadTrainsetFromFile('gdns/result_data_gdns_17_improved.csv')  # CSV 파일에서 데이터 로드
    x_trainset, x_testset = trainer.preprocessingXset(x_train, x_test, 'x_fited_scaler.v2.npz')  # 입력 데이터 정규화
    y_trainset, y_testset = trainer.preprocessingYset(y_train, y_test)  # 출력 데이터 분리
    
    # ===== 3단계: 3개 모델 설정 =====
    models = trainer.setupModels()  # [업체투찰률모델, 예가투찰률모델, 참여업체수모델]
    results = []  # 예측 결과를 저장할 리스트
    
    # ===== 4단계: 각 모델 훈련 및 저장 =====
    for i, model in enumerate(models):
        trainer.trainnng(model, x_trainset, y_trainset[i])  # 모델 훈련
        trainer.saveModel(model, f'mlpregr.model{i+1}.v0.1.1.npz')  # 모델 저장
        result = trainer.predict(model, x_testset)  # 테스트 데이터로 예측
        print(f"모델{i+1} 예측 결과 (처음 50개):")
        print(result[:50])
        print("="*80)
        results.append(result)  # 예측 결과 저장
    
    # 훈련된 모델들을 trainer 객체에 저장 (성능 측정용)
    trainer.model1 = models[0]
    trainer.model2 = models[1] 
    trainer.model3 = models[2]
        
    # ===== 5단계: 결과 정리 및 저장 =====
    print("="*80)
    print("📊 예측 결과를 데이터프레임으로 정리 중...")
    df_result = trainer.mergeResultset(results)  # 예측 결과를 데이터프레임으로 정리
    print(f"✅ 데이터프레임 생성 완료: {len(df_result)}행 x {len(df_result.columns)}열")
    
    print("="*80)
    print("💾 엑셀 파일로 결과 저장 중...")
    trainer.saveResultToXls(df_result, trainer.xlxs_dir)  # 엑셀 파일로 결과 저장

    # ===== 6단계: 모델 성능 측정 =====
    print("="*80)
    print("🔍 훈련된 모델 성능 측정 중...")
    print("="*80)
    
    # 훈련된 모델들로 성능 측정
    from sklearn.metrics import mean_squared_error, r2_score, mean_absolute_error
    
    models = [trainer.model1, trainer.model2, trainer.model3]
    model_names = ['업체투찰률', '예가투찰률', '참여업체수']
    y_tests = [y_testset[0], y_testset[1], y_testset[2]]
    
    performance_results = {}
    
    for i, (model, name, y_test) in enumerate(zip(models, model_names, y_tests)):
        try:
            # 훈련된 모델로 예측
            y_pred = model.predict(x_testset)
            
            # 성능 지표 계산
            mse = mean_squared_error(y_test, y_pred)
            r2 = r2_score(y_test, y_pred)
            mae = mean_absolute_error(y_test, y_pred)
            rmse = np.sqrt(mse)
            
            performance_results[name] = {
                'MSE': mse,
                'R2': r2,
                'MAE': mae,
                'RMSE': rmse
            }
            
            print(f"\n{name} 모델 성능:")
            print(f"  MSE:  {mse:.6f}")
            print(f"  R2:   {r2:.6f}")
            print(f"  MAE:  {mae:.6f}")
            print(f"  RMSE: {rmse:.6f}")
            
            # 데이터 분포 확인
            print(f"  실제값 범위: {y_test.min():.6f} ~ {y_test.max():.6f}")
            print(f"  예측값 범위: {y_pred.min():.6f} ~ {y_pred.max():.6f}")
            
            # 성능 해석
            if r2 > 0.8:
                r2_interpretation = "매우 좋음"
            elif r2 > 0.6:
                r2_interpretation = "좋음"
            elif r2 > 0.4:
                r2_interpretation = "보통"
            elif r2 > 0:
                r2_interpretation = "개선 필요"
            else:
                r2_interpretation = "매우 나쁨"
            
            print(f"  해석: R2 = {r2:.3f} ({r2_interpretation})")
            
        except Exception as e:
            print(f"ERROR: {name} 모델 성능 측정 오류: {e}")
    
    # 전체 성능 요약
    if performance_results:
        print(f"\n" + "="*80)
        print("전체 성능 요약")
        print("="*80)
        
        for name, metrics in performance_results.items():
            print(f"{name:>12}: R2={metrics['R2']:.3f}, RMSE={metrics['RMSE']:.3f}")
        
        # 평균 성능 계산
        avg_r2 = np.mean([metrics['R2'] for metrics in performance_results.values()])
        avg_rmse = np.mean([metrics['RMSE'] for metrics in performance_results.values()])
        
        print(f"\n평균 성능:")
        print(f"  평균 R2:   {avg_r2:.3f}")
        print(f"  평균 RMSE: {avg_rmse:.3f}")
        
        # 성능 기준선 설정
        print(f"\n성능 기준선:")
        if avg_r2 > 0.7:
            print("  현재 성능: 우수 (R2 > 0.7)")
            print("  개선 목표: R2 > 0.8")
        elif avg_r2 > 0.5:
            print("  현재 성능: 보통 (R2 > 0.5)")
            print("  개선 목표: R2 > 0.7")
        elif avg_r2 > 0:
            print("  현재 성능: 개선 필요 (R2 > 0)")
            print("  개선 목표: R2 > 0.5")
        else:
            print("  현재 성능: 매우 나쁨 (R2 < 0)")
            print("  개선 목표: R2 > 0.3")
    trainer.close()  # 훈련 과정 마무리

    
if __name__ == "__main__":
    """
    스크립트가 직접 실행될 때만 Main() 함수를 호출
    
    설명:
    - 이 파일이 직접 실행될 때만 머신러닝 훈련이 시작됨
    - 다른 파일에서 import할 때는 실행되지 않음
    """
    import sys
    
    # 명령행 인수 확인
    if len(sys.argv) > 1 and sys.argv[1] == "test":
        # 성능 테스트 실행
        print("모델 성능 테스트를 실행합니다...")
        test_model_performance()
    else:
        # 입찰 유형 확인 및 훈련 실행
        bid_type = 'auto'  # 기본값
        if len(sys.argv) > 1:
            bid_type = sys.argv[1]
            if bid_type not in ['cst', 'mtrl', 'gdns', 'auto']:
                print(f"⚠️  잘못된 입찰 유형: {bid_type}")
                print("사용 가능한 유형: cst, mtrl, gdns, auto")
                bid_type = 'auto'
        
        print(f"모델 훈련을 실행합니다... (입찰 유형: {bid_type})")
        Main(bid_type=bid_type)