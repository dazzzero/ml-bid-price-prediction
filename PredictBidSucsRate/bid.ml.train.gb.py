# -*- coding: utf-8 -*-
"""
입찰 성공률 예측 시스템 - 그래디언트 부스팅 모델 훈련 스크립트
Created on Mon Dec  2 10:14:33 2024

@author: user

이 파일의 목적:
- 조달청 입찰 데이터를 사용하여 머신러닝 모델을 훈련시키는 스크립트
- 3개의 그래디언트 부스팅 모델을 훈련 (업체투찰률, 예가투찰률, 참여업체수 예측)
- XGBoost, LightGBM, GradientBoostingRegressor 지원
- 훈련된 모델과 전처리 도구들을 파일로 저장
- 텍스트 데이터를 TF-IDF 방식으로 벡터화하여 숫자로 변환
"""

# ===== 필요한 라이브러리들 import =====
import os  # 파일 경로 조작을 위한 라이브러리
import joblib  # 머신러닝 모델을 파일로 저장/불러오기 위한 라이브러리
import numpy as np  # 수치 계산을 위한 라이브러리 (행렬, 배열 연산)
import random as rnd  # 랜덤 숫자 생성

from time import time  # 실행 시간 측정을 위한 라이브러리
import pandas as pd  # 데이터 분석을 위한 라이브러리 (엑셀과 비슷한 기능)

# 엑셀 파일 처리를 위한 라이브러리 (통계 추가용)
try:
    import openpyxl
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("⚠️  openpyxl이 설치되지 않았습니다. 엑셀 통계 기능을 사용하려면 'pip install openpyxl'을 실행하세요.")

# 한국어 자연어 처리를 위한 라이브러리
from kiwipiepy import Kiwi  # 한국어 형태소 분석기 (단어를 쪼개는 도구)

# 텍스트를 숫자로 변환하는 도구들
from sklearn.feature_extraction.text import TfidfVectorizer  # 텍스트를 숫자로 변환
from scipy.sparse import csr_matrix  # 메모리 효율적인 행렬 저장 방식

# 머신러닝 모델과 전처리 도구들
from sklearn.ensemble import GradientBoostingRegressor  # 그래디언트 부스팅 회귀 모델
from sklearn.preprocessing import StandardScaler  # 데이터를 정규화하는 도구 (0~1 사이로 맞춤)

# 고성능 그래디언트 부스팅 라이브러리들
try:
    import xgboost as xgb  # XGBoost (Extreme Gradient Boosting)
    XGBOOST_AVAILABLE = True
except ImportError:
    XGBOOST_AVAILABLE = False
    print("⚠️  XGBoost가 설치되지 않았습니다. 'pip install xgboost'를 실행하세요.")

try:
    import lightgbm as lgb  # LightGBM (Light Gradient Boosting Machine)
    LIGHTGBM_AVAILABLE = True
except ImportError:
    LIGHTGBM_AVAILABLE = False
    print("⚠️  LightGBM이 설치되지 않았습니다. 'pip install lightgbm'을 실행하세요.")

# 데이터 분할 도구
from sklearn.model_selection import train_test_split  # 훈련 데이터와 테스트 데이터로 분할


class KiwiTokenizer():
    """
    한국어 텍스트를 처리하는 클래스 (predict.py와 동일)
    
    이 클래스의 역할:
    1. 한국어 문장을 단어 단위로 쪼개기 (형태소 분석)
    2. 의미있는 단어만 추출하기 (명사, 형용사 등)
    3. 불필요한 문자 제거하기 (괄호, 특수문자 등)
    
    예시: "서울시청 건물 신축공사" → ["서울시청", "건물", "신축", "공사"]
    """
    
    def __init__(self, saved_filenm):
        """
        KiwiTokenizer 초기화 함수
        
        Args:
            saved_filenm (str): 저장된 사전 파일명 (이전에 학습한 단어사전을 불러올 때 사용)
        """
        # 랜덤 번호 생성 (파일명에 사용)
        self.rnd_num = rnd.randint(100, 999)
        
        # 현재 작업 디렉토리 경로 설정
        self.cur_dir = os.getcwd()  # 현재 폴더 경로 가져오기
        self.data_dir = self.cur_dir+'\\data\\'  # 데이터 폴더 경로 (입찰 데이터가 있는 곳)
        self.save_dir = self.cur_dir+'\\res\\'  # 결과 저장 폴더 (모델 파일들이 있는 곳)
        self.save_filename = saved_filenm  # 저장할 파일명
        
        # Kiwi 형태소 분석기 초기화
        self.kiwi = None
        self.kiwi = self.CreateKiwi(saved_filenm)  # Kiwi 객체 생성
    
    def save(self, filename):
        joblib.dump(self.kiwi._user_values, self.save_dir+filename)
        return self
    
    def load(self, filename):
        o = joblib.load(self.save_dir+filename)
        return o
        
    def loadDictonary(self, filename, max_rows: int = 30000, log_every: int = 2000):
        """
        대용량 사용자 사전 구축 최적화: 상한선(cap)과 진행 로그를 추가해 장시간 블로킹을 방지한다.
        """
        full_path = self.data_dir+filename
        if not os.path.exists(full_path):
            print(f"⚠️  사용자 사전 파일이 없습니다: {full_path}")
            return

        try:
            self.data = pd.read_csv(full_path, nrows=max_rows)
        except Exception as e:
            print(f"⚠️  사용자 사전 로드 실패: {e}")
            return

        if "단어" not in self.data.columns:
            print("⚠️  '단어' 컬럼이 없어 사용자 사전을 구축할 수 없습니다.")
            return

        print(f"📘 사용자 사전 단어 추가 시작 (최대 {len(self.data)}건)")
        for i, row in enumerate(self.data.itertuples(index=False)):
            try:
                self.kiwi.add_user_word(getattr(row, "단어"), 'NNP')
            except KeyboardInterrupt:
                print("⏹️  사용자 중단으로 사전 구축을 중지합니다.")
                break
            except Exception:
                # 개별 단어 추가 실패는 무시하고 진행
                pass
            if (i+1) % log_every == 0:
                print(f"   - 진행: {i+1} / {len(self.data)}")
        print("✅ 사용자 사전 단어 추가 완료")
    
    def CreateKiwi(self, saved_filenm):
        o = None
        if(self.kiwi is None):
            if(saved_filenm is None):
                o = Kiwi(num_workers=8)
            else:
                o = Kiwi(num_workers=8)
                # 저장된 사용자 사전을 로드할 때는 파일명만 전달 (load 내부에서 경로 결합)
                o._user_values = self.load(saved_filenm)
                #o.load_user_dictionary(self.save_dir+saved_filenm)
        
        return o
            
    def cleared_line(self, line):
        # numpy.float64나 NaN 값 처리
        if pd.isna(line) or line is None:
            line = ''
        else:
            line = str(line).lower().replace('(', ' ').replace(')', ' ').replace('n/a', '')
        
        nm_words = []
        tokens = self.kiwi.tokenize(line)
        for token in tokens:
            if token.tag in ['MM', 'NNG', 'NNB', 'NNP', 'SL', 'XPN', 'MAG', 'SN', 'SO', 'W_SERIAL']:
                nm_words.append(token.form)
                
        return ' '.join(nm_words)
    
    def cleared_lines_from(self, orglines):
        lines = []
        for line in orglines:
            lines.append(self.cleared_line(line))
        
        return lines      
    
    def get_key(self, voca, val):
      
        for key, value in voca.items():
            if val == value:
                return key
    
        return "key doesn't exist"
    
    def conv_words(self, voca, vals):
        ret = []
        for val in vals:
            for key, value in voca.items():
                if val == value:
                    ret.append(key)
            
            
        return " ".join(ret)
    
    
    def nn_only(self, orglines):
        lines = []
        for key in orglines:
            # numpy.float64나 NaN 값 처리
            if pd.isna(key) or key is None:
                key = ''
            else:
                key = str(key).lower().replace('(', ' ').replace(')', ' ').replace('n/a', '')
            
            nm_words = []
            tokens = self.kiwi.tokenize(key)
            #print(key, tokens)
            for token in tokens:
                if token.tag in ['MM', 'NNG', 'NNB', 'NNP', 'SL', 'XPN', 'MAG', 'SN', 'SO', 'W_SERIAL']:
                    nm_words.append(token.form)
            lines.append(' '.join(nm_words))
            
        return lines


class KiwiVectorizer():
    def __init__(self):
        self.rnd_num = rnd.randint(100, 999)
        
        self.cur_dir = os.getcwd()
        self.save_dir = self.cur_dir+'\\res\\'
        
        # 메모리 효율성을 위해 파라미터 최적화
        self.vect = TfidfVectorizer(
            ngram_range=(1, 1), 
            token_pattern='(?u)\\b\\w+\\b',
            max_features=5000,  # 최대 특성 수 제한
            min_df=2,           # 최소 문서 빈도 (2회 이상 나타나는 단어만)
            max_df=0.95,        # 최대 문서 빈도 (95% 이상 문서에 나타나는 단어 제외)
            sublinear_tf=True   # TF 스케일링으로 메모리 절약
        )
    
    def load(self, filename):
        voca = joblib.load(self.save_dir + filename)
        self.vect.vocabulary_ = voca["vocabulary"] 
        self.vect.idf_ = voca["idf"]
    
    def save(self, filename):
        joblib.dump({'vocabulary':self.vect.vocabulary_, 'idf':self.vect.idf_}, self.save_dir + filename)
        print("단어사전 파일("+filename+")로 저장합니다. ")
    
    def fit(self, lines):
        return self.vect.fit(lines)
    
    def transform(self, lines):
        # 메모리 효율성을 위해 toarray() 없이 CSR 행렬 그대로 반환
        return self.vect.transform(lines)
    
    def scores(self, lines):
        return self.toValues(self.transform(lines))


    def toValues(self, csrmat:csr_matrix):
        lst = []
        safty_sz = len(csrmat.indptr)-1
        for i, n in enumerate(csrmat.indptr):
            sumval = 0
            if i<safty_sz:
                indiceVal = csrmat.indices[csrmat.indptr[i]:csrmat.indptr[i+1]]
                dataVal = csrmat.data[csrmat.indptr[i]:csrmat.indptr[i+1]]
                for j, va1 in enumerate(indiceVal):
                    sumval += va1 * dataVal[j]
                
                lst.append(sumval)
        return lst


class BidLowerMarginRateTrain():
    """
    입찰 투찰률 예측 모델을 훈련시키는 메인 클래스
    
    이 클래스의 역할:
    1. 조달청 입찰 데이터를 불러와서 전처리
    2. 텍스트 데이터(키워드, 기관명, 지역)를 숫자로 변환
    3. 3개의 그래디언트 부스팅 모델을 훈련
       - 모델1: 업체 투찰률 예측 (XGBoost)
       - 모델2: 예가 투찰률 예측 (LightGBM)  
       - 모델3: 참여 업체 수 예측 (GradientBoostingRegressor)
    4. 훈련된 모델과 전처리 도구들을 파일로 저장
    5. 예측 결과를 엑셀 파일로 출력
    
    머신러닝 과정:
    1. 데이터 로드 → 2. 텍스트 벡터화 → 3. 데이터 정규화 → 4. 모델 훈련 → 5. 모델 저장
    """
    
    def __init__(self):
        """
        BidLowerMarginRateTrain 초기화 함수
        - 디렉토리 경로 설정
        - 데이터 컬럼 정의
        - 텍스트 처리 도구들 초기화
        """
        # 랜덤 번호 생성 (파일명에 사용)
        self.rnd_num = rnd.randint(100, 999)
        
        # 디렉토리 경로 설정
        self.cur_dir = os.getcwd()  # 현재 작업 디렉토리
        self.data_dir = self.cur_dir+'\\data\\'  # 데이터 폴더 (입찰 데이터가 있는 곳)
        self.save_dir = self.cur_dir+'\\res\\'  # 결과 저장 폴더 (모델 파일들이 저장될 곳)
        
        # 결과 엑셀 파일 경로 설정 (나중에 데이터 크기와 테스트 비율에 따라 생성)
        self.excel_file_nm = None  # 결과 엑셀 파일명 (나중에 설정)
        self.xlxs_dir = None  # 엑셀 파일 전체 경로 (나중에 설정) 
        
       
        # ===== 데이터 컬럼 정의 =====
        # CSV 파일에서 읽어올 컬럼들의 이름 정의
        self.cvs_columns = ['기초금액', '낙찰하한률', '참여업체수', 
                            '낙찰금액', '업체투찰률', '예가투찰률', '투찰률오차', 
                            '간접비', '순공사원가', '입찰번호', '입찰차수', 
                            '예정금액', '낙찰하한가', 
                            '면허제한코드','공고기관코드','주공종명', 
                            '공고기관명', '공고기관점수',
                            '공사지역', '공사지역점수',
                            '키워드', '키워드점수']
        
        # 각 컬럼의 데이터 타입 정의 (pandas가 올바르게 데이터를 읽기 위해)
        self.cvs_columns_type = {
                            '기초금액':'float64', '낙찰하한률':'float64', '참여업체수':'float64', 
                            '낙찰금액':'int64', '업체투찰률':'float64', '예가투찰률':'float64', '투찰률오차':'float64', 
                            '간접비':'int64', '순공사원가':'int64', '입찰번호':'str', '입찰차수':'int64', 
                            '예정금액':'int64', '낙찰하한가':'int64',
                            '면허제한코드':'str','공고기관코드':'str','주공종명':'str', 
                            '공고기관명':'str', '공고기관점수':'float64',
                            '공사지역':'str', '공사지역점수':'float64',
                            '키워드':'str', '키워드점수':'float64'
                            }
        
        # ===== 결과 데이터 컬럼 정의 =====
        # 예측 결과를 저장할 엑셀 파일의 컬럼들 정의 (원본 데이터 + 예측 결과)
        self.result_columns = ['입찰번호', '입찰차수', '기초금액', 
                               '낙찰하한률', '참여업체수', '간접비', '순공사원가', 
                               '면허제한코드', '공고기관코드',
                               '공고기관명', '공고기관점수',
                               '공사지역', '공사지역점수',                               
                               '키워드', '키워드점수',
                               '업체투찰률', '예가투찰률', '투찰률오차', 
                               '예정금액', '낙찰하한가', '낙찰금액', 
                               '업체투찰률예측', '예가투찰률예측', '참여업체수예측', '예정금액예측',
                               '낙찰금액(업체투찰률) 예측', 'A값여부', '결과1', 
                               '예정금액(예가투찰률) 예측', '예정금액*낙찰하한율', '결과2']
        
        # 결과 컬럼들의 데이터 타입 정의
        self.result_columns_type = {
                            '입찰번호':'str', '입찰차수':'int64','기초금액':'float64',
                            '낙찰하한률':'float64', '참여업체수':'float64', '간접비':'int64', '순공사원가':'int64',  
                            '면허제한코드':'float64', '공고기관코드':'float64',
                            '공고기관명':'str', '공고기관점수':'float64',
                            '공사지역':'str', '공사지역점수':'float64',                            
                            '키워드':'str', '키워드점수':'float64',
                            '업체투찰률':'float64', '예가투찰률':'float64', '투찰률오차':'float64', 
                            '예정금액':'int64', '낙찰하한가':'int64', '낙찰금액':'int64', 
                            '업체투찰률예측':'float64', '예가투찰률예측':'float64', '참여업체수예측':'float64', '예정금액예측':'int64',
                            '낙찰금액(업체투찰률) 예측':'float64', 'A값여부':'str', '결과1':'str',
                            '예정금액(예가투찰률) 예측':'float64', '예정금액*낙찰하한율':'float64', '결과2':'str'
                            }

        
        # ===== 텍스트 처리 도구들 초기화 =====
        # 한국어 형태소 분석기 초기화 (캐시 사용: 존재 시 로드, 없으면 생성 후 저장)
        tok_cache_filename = 'gb.tokenizer.v0.1.1.npz'
        if os.path.exists(self.save_dir + tok_cache_filename):
            print("✅ 사용자 사전 캐시를 로드합니다.")
            self.tokenizer = KiwiTokenizer(tok_cache_filename)
        else:
            print("📘 사용자 사전을 최초 구축합니다 (최초 1회만 수행).")
            self.tokenizer = KiwiTokenizer(None)
            # 대용량 구축 시 장시간이 걸릴 수 있어 일부만 로드하도록 최적화
            self.tokenizer.loadDictonary('표준국어대사전.NNP.csv')  # 표준국어대사전을 로드하여 정확한 단어 인식
            self.tokenizer.save(tok_cache_filename)
        
        # TF-IDF 벡터화기 초기화 (텍스트를 숫자로 변환하는 도구)
        self.vectorizer = KiwiVectorizer()

    def generateExcelFileName(self, data_size, test_ratio):
        """
        데이터 크기와 테스트 비율에 따라 엑셀 파일명을 생성하는 함수
        
        Args:
            data_size (int): 전체 데이터 크기
            test_ratio (float): 테스트 데이터 비율 (0.0 ~ 1.0)
            
        Returns:
            str: 생성된 파일명 (result-yyMMddMMss-데이터갯수-테스트비율.xlsx)
            
        파일명 규칙:
        - result-yyMMddMMss-데이터갯수-테스트비율.xlsx
        - 예: result-2412011430-5-8515.xlsx (2024년 12월 1일 14시 30분, 5만개 데이터, 85:15 비율)
        """
        from datetime import datetime
        
        # 현재 시간을 yyMMddMMss 형식으로 변환
        now = datetime.now()
        time_str = now.strftime("%y%m%d%H%M")
        
        # 데이터 크기를 만개 단위로 변환
        if data_size < 50000:
            data_unit = 1  # 5만개 미만은 1로 표시
        elif data_size < 100000:
            data_unit = 5  # 5만~10만개는 5로 표시
        elif data_size < 300000:
            data_unit = 10  # 10만~30만개는 10으로 표시
        else:
            data_unit = 30  # 30만개 이상은 30으로 표시
        
        # 테스트 비율을 정수로 변환 (예: 0.15 -> 15, 0.2 -> 20)
        train_ratio = int((1 - test_ratio) * 100)
        test_ratio_int = int(test_ratio * 100)
        ratio_str = f"{train_ratio}{test_ratio_int:02d}"  # 8515, 8020 등
        
        # 파일명 생성
        filename = f"result-{time_str}-{data_unit}-{ratio_str}.xlsx"
        
        return filename

    
    def make_result_dataframe2(self, xx_test, result1, result2, result3):
        df_rst = pd.DataFrame(columns = self.result_columns)
        df_rst.astype(self.result_columns_type)        
        
        selected_cols = ['입찰번호', '입찰차수', 
                         '기초금액', '낙찰하한률', '참여업체수', '간접비', '순공사원가', 
                         '면허제한코드', '공고기관코드', 
                         '키워드', '키워드점수', 
                         '공고기관명', '공고기관점수',
                         '공사지역', '공사지역점수',                             
                         '업체투찰률', '예가투찰률', '투찰률오차', 
                         '예정금액', '낙찰하한가', '낙찰금액']
        
        df_xx_test = pd.DataFrame(xx_test, columns = self.cvs_columns)
        df_test = pd.DataFrame(df_xx_test, columns = selected_cols)
        
        print("예측결과값 입력 시작")
        
        for i, v in enumerate(selected_cols):
            df_rst[v] = df_test[v]
                
        df_rst["업체투찰률예측"] = result1
        df_rst["예가투찰률예측"] = result2
        df_rst["참여업체수예측"] = result3
        
        # 예가투찰률 예측값으로부터 예정금액 역산 계산
        # 예가투찰률 = (예정금액 / 기초금액) * 낙찰하한률
        # 따라서 예정금액 = (예가투찰률 * 기초금액) / 낙찰하한률
        df_rst["예정금액예측"] = (df_rst["예가투찰률예측"] * df_rst["기초금액"]) / df_rst["낙찰하한률"]
        
        # ===== 새로운 컬럼들 계산 =====
        # 낙찰금액(업체투찰률) 예측 = 업체투찰률예측 * 기초금액
        df_rst["낙찰금액(업체투찰률) 예측"] = df_rst["업체투찰률예측"] * df_rst["기초금액"]
        
        # A값여부: 업체투찰률예측이 0.8 이상이면 'O', 아니면 빈 문자열
        df_rst["A값여부"] = df_rst["업체투찰률예측"].apply(lambda x: 'O' if x >= 0.8 else '')
        
        # 결과1: 낙찰금액(업체투찰률) 예측이 낙찰하한가보다 작으면 "낙찰하한선미달", 
        # 낙찰하한가 이상이고 낙찰금액 미만이면 "낙찰", 아니면 "-"
        def calculate_result1(row):
            predicted_amount = row["낙찰금액(업체투찰률) 예측"]
            min_bid = row["낙찰하한가"]
            actual_amount = row["낙찰금액"]
            
            if predicted_amount < min_bid:
                return "낙찰하한선미달"
            elif predicted_amount >= min_bid and predicted_amount < actual_amount:
                return "낙찰"
            else:
                return "-"
        
        df_rst["결과1"] = df_rst.apply(calculate_result1, axis=1)
        
        # 예정금액(예가투찰률) 예측 = 예가투찰률예측 / 낙찰하한률 * 기초금액
        df_rst["예정금액(예가투찰률) 예측"] = (df_rst["예가투찰률예측"] / df_rst["낙찰하한률"]) * df_rst["기초금액"]
        
        # 예정금액*낙찰하한율 = 예정금액(예가투찰률) 예측 * 낙찰하한률
        df_rst["예정금액*낙찰하한율"] = df_rst["예정금액(예가투찰률) 예측"] * df_rst["낙찰하한률"]
        
        # 결과2: 예정금액*낙찰하한율이 낙찰하한가보다 작으면 "낙찰하한선미달",
        # 낙찰하한가 이상이고 낙찰금액 미만이면 "낙찰", 아니면 "-"
        def calculate_result2(row):
            predicted_amount = row["예정금액*낙찰하한율"]
            min_bid = row["낙찰하한가"]
            actual_amount = row["낙찰금액"]
            
            if predicted_amount < min_bid:
                return "낙찰하한선미달"
            elif predicted_amount >= min_bid and predicted_amount < actual_amount:
                return "낙찰"
            else:
                return "-"
        
        df_rst["결과2"] = df_rst.apply(calculate_result2, axis=1)
        
        print("예측결과테이블 작성완료")
         
        return df_rst


    
    def arrayToDataFrame(self, data, cols):
        df = pd.DataFrame(pd.DataFrame(data), columns=cols)
        return df
    

    def loadTrainsetFromFile(self, filename):
        """
        CSV 파일에서 입찰 데이터를 불러와서 전처리하는 함수
        
        Args:
            filename (str): 불러올 CSV 파일명
            
        Returns:
            tuple: (x_train, x_test, y_train, y_test) - 훈련/테스트 데이터
            
        처리 과정:
        1. CSV 파일 읽기
        2. 텍스트 데이터(키워드, 기관명, 지역)를 TF-IDF 점수로 변환
        3. 훈련 데이터와 테스트 데이터로 분할
        4. 필요한 컬럼만 선택하여 반환
        """
        print("원시 훈련데이타를 불러옵니다.")
        #'bid_v5_202412311021.csv'
        data = pd.read_csv(self.data_dir+filename)   # CSV 파일 읽기 (예: bid_250914.csv)
        
        
        print("총 데이타수: "+str(len(data))+'건')
        
        print("="*80)
        print("학습셋과 테스트셋 분리")
        print(f"전체 데이터: {len(data)}개")
        
        # ===== 데이터 크기에 따른 동적 테스트 비율 설정 =====
        data_size = len(data)
        if data_size < 20000:
            test_ratio = 0.2  # 20% - 2만개 미만
            ratio_desc = "80:20 (소규모 데이터)"
        elif data_size < 50000:
            test_ratio = 0.15  # 15% - 2만~5만개
            ratio_desc = "85:15 (중소규모 데이터)"
        elif data_size < 100000:
            test_ratio = 0.1   # 10% - 5만~10만개
            ratio_desc = "90:10 (중규모 데이터)"
        else:
            test_ratio = 0.05  # 5% - 10만개 이상
            ratio_desc = "95:5 (대규모 데이터)"
        
        train_count = int(data_size * (1 - test_ratio))
        test_count = data_size - train_count
        
        print(f"적용 비율: {ratio_desc}")
        print(f"훈련 데이터: {train_count:,}개 ({((1-test_ratio)*100):.0f}%)")
        print(f"테스트 데이터: {test_count:,}개 ({test_ratio*100:.0f}%)")
        print("="*80)
        
        # ===== 입력 데이터(X) 준비 =====
        # 필요한 컬럼들만 선택하여 데이터프레임 생성
        dataset_x = pd.DataFrame(data, columns = self.cvs_columns)
        
        # 텍스트 컬럼들을 먼저 문자열로 변환
        print("텍스트 컬럼을 문자열로 변환 중...")
        if '키워드' in dataset_x.columns:
            dataset_x['키워드'] = dataset_x['키워드'].fillna('').astype(str)
            dataset_x['키워드'] = dataset_x['키워드'].replace(['nan', 'NaN', 'None', 'null'], '')
        if '공고기관명' in dataset_x.columns:
            dataset_x['공고기관명'] = dataset_x['공고기관명'].fillna('').astype(str)
            dataset_x['공고기관명'] = dataset_x['공고기관명'].replace(['nan', 'NaN', 'None', 'null'], '')
        if '공사지역' in dataset_x.columns:
            dataset_x['공사지역'] = dataset_x['공사지역'].fillna('').astype(str)
            dataset_x['공사지역'] = dataset_x['공사지역'].replace(['nan', 'NaN', 'None', 'null'], '')
        
        # 문자열 컬럼들을 숫자로 변환
        print("문자열 컬럼을 숫자로 변환 중...")
        
        # 면허제한코드를 숫자로 변환 (문자열을 해시값으로 변환)
        if '면허제한코드' in dataset_x.columns:
            dataset_x['면허제한코드'] = dataset_x['면허제한코드'].astype(str).apply(lambda x: hash(x) % 1000000 if pd.notna(x) and x != 'nan' else 0)
        
        # 공고기관코드를 숫자로 변환
        if '공고기관코드' in dataset_x.columns:
            dataset_x['공고기관코드'] = dataset_x['공고기관코드'].astype(str).apply(lambda x: hash(x) % 1000000 if pd.notna(x) and x != 'nan' else 0)
        
        # 각 컬럼의 데이터 타입을 올바르게 설정
        dataset_x.astype(self.cvs_columns_type)
        
        
        # ===== 텍스트 데이터 처리 =====
        # 각 텍스트 컬럼을 형태소 분석하여 의미있는 단어만 추출
        print("키워드 컬럼 데이터 타입 및 샘플:")
        print(f"키워드 컬럼 타입: {dataset_x['키워드'].dtype}")
        print(f"키워드 샘플 (처음 10개): {dataset_x['키워드'].head(10).tolist()}")
        print(f"키워드 NaN 개수: {dataset_x['키워드'].isna().sum()}")
        
        print("\n공고기관명 컬럼 데이터 타입 및 샘플:")
        print(f"공고기관명 컬럼 타입: {dataset_x['공고기관명'].dtype}")
        print(f"공고기관명 샘플 (처음 10개): {dataset_x['공고기관명'].head(10).tolist()}")
        print(f"공고기관명 NaN 개수: {dataset_x['공고기관명'].isna().sum()}")
        print(f"공고기관명 고유값 개수: {dataset_x['공고기관명'].nunique()}")
        
        print("\n공사지역 컬럼 데이터 타입 및 샘플:")
        print(f"공사지역 컬럼 타입: {dataset_x['공사지역'].dtype}")
        print(f"공사지역 샘플 (처음 10개): {dataset_x['공사지역'].head(10).tolist()}")
        print(f"공사지역 NaN 개수: {dataset_x['공사지역'].isna().sum()}")
        
        lines = self.tokenizer.nn_only(np.squeeze(dataset_x["키워드"].tolist()))      # 키워드 처리
        lines2 = self.tokenizer.nn_only(np.squeeze(dataset_x["공고기관명"].tolist()))  # 공고기관명 처리
        lines3 = self.tokenizer.nn_only(np.squeeze(dataset_x["공사지역"].tolist()))    # 공사지역 처리
        
        # 처리 결과 미리보기 (처음 10개)
        print("키워드 처리 결과:")
        print(lines[:10])
        print("공고기관명 처리 결과:")
        print(lines2[:10])
        print("공사지역 처리 결과:")
        print(lines3[:10])
        
        # ===== TF-IDF 벡터화기 학습 =====
        # 모든 텍스트 데이터를 합쳐서 단어사전을 학습
        all_text = lines + lines2 + lines3
        print(f"전체 텍스트 샘플 수: {len(all_text)}")
        print(f"빈 문자열 개수: {sum(1 for text in all_text if text.strip() == '')}")
        
        # 빈 문자열이 아닌 텍스트만 사용하여 학습
        non_empty_text = [text for text in all_text if text.strip() != '']
        print(f"비어있지 않은 텍스트 수: {len(non_empty_text)}")
        
        if len(non_empty_text) > 0:
            self.vectorizer.fit(non_empty_text)
        else:
            # 모든 텍스트가 비어있다면 기본값으로 학습
            self.vectorizer.fit(['기본키워드'])
        
        # ===== 텍스트를 TF-IDF 점수로 변환 =====
        pts = self.vectorizer.scores(lines)      # 키워드 점수 계산
        pts2 = self.vectorizer.scores(lines2)    # 공고기관명 점수 계산
        pts3 = self.vectorizer.scores(lines3)    # 공사지역 점수 계산
        
        # 점수 결과 미리보기 (처음 10개)
        print("키워드 점수:")
        print(pts[:10])
        print("공고기관명 점수:")
        print(pts2[:10])
        print("공사지역 점수:")
        print(pts3[:10])
        
        # ===== 점수를 데이터프레임에 추가 =====
        dataset_x["키워드점수"] = pts        # 키워드 TF-IDF 점수 추가
        dataset_x["공고기관점수"] = pts2      # 공고기관명 TF-IDF 점수 추가
        dataset_x["공사지역점수"] = pts3      # 공사지역 TF-IDF 점수 추가
        
        # ===== 학습된 도구들을 파일로 저장 =====
        self.tokenizer.save("gb.tokenizer.v0.1.1.npz")    # 형태소 분석기 저장
        self.vectorizer.save("gb.vectorizer.v0.1.1.npz")  # TF-IDF 벡터화기 저장
        
        # ===== 출력 데이터(Y) 준비 =====
        # 예측할 대상 변수들 (업체투찰률, 예가투찰률, 참여업체수)
        dataset_y = pd.DataFrame(data, columns = ['업체투찰률', '예가투찰률', '참여업체수'])
        dataset_y.astype({'업체투찰률':'float64', '예가투찰률':'float64', '참여업체수':'int64'})
        # 주석처리된 이전 버전: '투찰률오차'도 포함했었음
        #dataset_y.astype({'업체투찰률':'float64', '예가투찰률':'float64', '투찰률오차':'float64'})

        # ===== 엑셀 파일명 생성 =====
        # 데이터 크기와 테스트 비율에 따라 파일명 생성
        self.excel_file_nm = self.generateExcelFileName(data_size, test_ratio)
        self.xlxs_dir = os.path.join(self.save_dir, self.excel_file_nm)
        
        print(f"📁 생성된 엑셀 파일명: {self.excel_file_nm}")
        print(f"📂 저장 경로: {self.xlxs_dir}")
        
        # ===== 훈련 데이터와 테스트 데이터로 분할 =====
        # 데이터 크기에 따라 동적으로 설정된 비율로 분할
        self.xx_train, self.xx_test, self.yy_train, self.yy_test = train_test_split(
                                                            dataset_x.to_numpy(),  # 입력 데이터 (X)
                                                            dataset_y.to_numpy(),  # 출력 데이터 (Y)
                                                            test_size=test_ratio,  # 동적으로 설정된 테스트 데이터 비율
                                                            random_state=self.rnd_num  # 랜덤 시드 (재현 가능한 결과)
                                                            )
        
        # ===== 필요한 컬럼만 선택 =====
        # 인덱스 [0,1,2,7,8,13,17,19,21]에 해당하는 컬럼들만 사용
        # 0:기초금액, 1:낙찰하한률, 2:참여업체수, 7:간접비, 8:순공사원가, 
        # 13:면허제한코드, 17:공고기관점수, 19:공사지역점수, 21:키워드점수
        x_train = (self.arrayToDataFrame(self.xx_train, [0,1,2,7,8,13,17,19,21])).to_numpy()
        x_test = (self.arrayToDataFrame(self.xx_test, [0,1,2,7,8,13,17,19,21])).to_numpy()
        
        return x_train, x_test, self.yy_train, self.yy_test
        
    def preprocessingXset(self, x_train, x_test, scalerSaveName):
        """
        입력 데이터를 정규화하는 함수
        
        Args:
            x_train (numpy.array): 훈련용 입력 데이터
            x_test (numpy.array): 테스트용 입력 데이터
            scalerSaveName (str): 정규화 도구를 저장할 파일명
            
        Returns:
            tuple: (x_trainset, x_testset) - 정규화된 훈련/테스트 데이터
            
        설명:
        - StandardScaler를 사용하여 데이터를 평균 0, 표준편차 1로 정규화
        - 머신러닝 모델이 더 잘 학습할 수 있도록 데이터 스케일을 맞춤
        - 정규화 도구를 파일로 저장하여 나중에 예측 시에도 동일하게 적용
        """
        print("="*80)
        print("데이타셋 정규화중...")
        
        # ===== 정규화 도구 선택 =====
        self.scaler = StandardScaler()  # 표준 정규화 (평균 0, 표준편차 1)
        # 다른 정규화 방법들 (주석처리)
        #scaler = MinMaxScaler()    # 최소-최대 정규화 (0~1 범위)
        #scaler = RobustScaler()    # 로버스트 정규화 (이상치에 강함)
        
        # ===== 훈련 데이터로 정규화 도구 학습 =====
        x_trainset = self.scaler.fit_transform(x_train).tolist()  # 훈련 데이터 정규화
        
        # ===== 테스트 데이터 정규화 =====
        x_testset = (self.scaler.transform(x_test)).tolist()  # 테스트 데이터 정규화
        
        # ===== 정규화 도구를 파일로 저장 =====
        #'gb_scaler.v2.npz'
        joblib.dump(self.scaler, self.save_dir+scalerSaveName)  # 나중에 예측 시 사용할 수 있도록 저장
        
        return x_trainset, x_testset
    
    def preprocessingYset(self, yy_train, yy_test):
        """
        출력 데이터를 3개의 모델용으로 분리하는 함수
        
        Args:
            yy_train (numpy.array): 훈련용 출력 데이터 [업체투찰률, 예가투찰률, 참여업체수]
            yy_test (numpy.array): 테스트용 출력 데이터 [업체투찰률, 예가투찰률, 참여업체수]
            
        Returns:
            tuple: (y_trainset, y_testset) - 각각 3개 모델용 데이터 리스트
            
        설명:
        - 각 모델이 예측할 대상 변수를 분리
        - 모델1: 업체투찰률, 모델2: 예가투찰률, 모델3: 참여업체수
        """
        # ===== 훈련 데이터를 3개 모델용으로 분리 =====
        y_train1 = np.squeeze( (self.arrayToDataFrame(yy_train, [0])).to_numpy() )  # 업체투찰률
        y_train2 = np.squeeze( (self.arrayToDataFrame(yy_train, [1])).to_numpy() )  # 예가투찰률
        y_train3 = np.squeeze( (self.arrayToDataFrame(yy_train, [2])).to_numpy() )  # 참여업체수
        
        # ===== 테스트 데이터를 3개 모델용으로 분리 =====
        y_test1 = np.squeeze( (self.arrayToDataFrame(yy_test, [0])).to_numpy() )    # 업체투찰률
        y_test2 = np.squeeze( (self.arrayToDataFrame(yy_test, [1])).to_numpy() )    # 예가투찰률
        y_test3 = np.squeeze( (self.arrayToDataFrame(yy_test, [2])).to_numpy() )    # 참여업체수
        
        return [y_train1, y_train2, y_train3], [y_test1, y_test2, y_test3]
        
    
    
    def setupModels(self):
        """
        3개의 그래디언트 부스팅 모델을 설정하는 함수
        
        Returns:
            list: [model1, model2, model3] - 설정된 3개의 모델 리스트
            
        모델 설명:
        - model1: 업체 투찰률 예측 모델 (XGBoost)
        - model2: 예가 투찰률 예측 모델 (LightGBM)  
        - model3: 참여 업체 수 예측 모델 (GradientBoostingRegressor)
        
        그래디언트 부스팅 파라미터 설명:
        - n_estimators: 부스팅 단계 수 (더 많을수록 정확하지만 느림)
        - max_depth: 나무의 최대 깊이
        - learning_rate: 학습률 (작을수록 안정적이지만 느림)
        - subsample: 각 부스팅 단계에서 사용할 샘플 비율
        - random_state: 랜덤 시드 (재현 가능한 결과)
        """
        print("="*80)
        print("그래디언트 부스팅 모델 설정 시작...")
        print("="*80)
        
        self.t0 = time()  # 시작 시간 기록

        # ===== 3개의 그래디언트 부스팅 모델 설정 =====
        
        # 모델1: 업체 투찰률 예측 모델 (XGBoost)
        if XGBOOST_AVAILABLE:
            model1 = xgb.XGBRegressor(
                                n_estimators = 200,        # 부스팅 단계 200개
                                max_depth = 6,             # 최대 깊이 6
                                learning_rate = 0.1,       # 학습률 0.1
                                subsample = 0.8,           # 80% 샘플 사용
                                colsample_bytree = 0.8,    # 80% 특성 사용
                                random_state = 1,          # 랜덤 시드
                                n_jobs = -1,               # 모든 CPU 사용
                                verbosity = 1              # 진행상황 출력
                                )
            print("✅ XGBoost 모델1 설정 완료")
        else:
            # XGBoost가 없으면 GradientBoostingRegressor 사용
            model1 = GradientBoostingRegressor(
                                n_estimators = 200,
                                max_depth = 6,
                                learning_rate = 0.1,
                                subsample = 0.8,
                                random_state = 1
                                )
            print("⚠️  XGBoost 없음 - GradientBoostingRegressor 모델1 사용")
        
        # 모델2: 예가 투찰률 예측 모델 (LightGBM)
        if LIGHTGBM_AVAILABLE:
            model2 = lgb.LGBMRegressor(
                                n_estimators = 200,        # 부스팅 단계 200개
                                max_depth = 6,             # 최대 깊이 6
                                learning_rate = 0.1,       # 학습률 0.1
                                subsample = 0.8,           # 80% 샘플 사용
                                colsample_bytree = 0.8,    # 80% 특성 사용
                                random_state = 1,          # 랜덤 시드
                                n_jobs = -1,               # 모든 CPU 사용
                                verbosity = 1              # 진행상황 출력
                                )
            print("✅ LightGBM 모델2 설정 완료")
        else:
            # LightGBM이 없으면 GradientBoostingRegressor 사용
            model2 = GradientBoostingRegressor(
                                n_estimators = 200,
                                max_depth = 6,
                                learning_rate = 0.1,
                                subsample = 0.8,
                                random_state = 1
                                )
            print("⚠️  LightGBM 없음 - GradientBoostingRegressor 모델2 사용")
        
        # 모델3: 참여 업체 수 예측 모델 (GradientBoostingRegressor)
        model3 = GradientBoostingRegressor(
                            n_estimators = 300,        # 부스팅 단계 300개 (더 많음)
                            max_depth = 8,             # 최대 깊이 8 (더 깊음)
                            learning_rate = 0.05,      # 학습률 0.05 (더 작음)
                            subsample = 0.9,           # 90% 샘플 사용
                            random_state = 1,          # 랜덤 시드
                            verbose = 1                # 진행상황 출력
                            )
        print("✅ GradientBoostingRegressor 모델3 설정 완료")
        
        return [model1, model2, model3]
    
    def trainnng(self, model, x_trainset, y_trainset):
        """
        머신러닝 모델을 훈련시키는 함수
        
        Args:
            model: 훈련시킬 그래디언트 부스팅 모델
            x_trainset (list): 훈련용 입력 데이터
            y_trainset (list): 훈련용 출력 데이터
            
        설명:
        - 모델이 입력 데이터를 보고 출력 데이터를 예측하도록 학습
        - 그래디언트 부스팅이 순차적으로 오차를 줄여가며 예측 정확도를 높임
        """
        # 훈련 데이터 미리보기 (처음 50개)
        print("훈련 데이터 미리보기:")
        print(y_trainset[:50])
        
        # ===== 모델 훈련 실행 =====
        model.fit(x_trainset, y_trainset)  # 모델이 데이터를 학습

        print("-"*80)
        print("그래디언트 부스팅 모델로 학습을 완료하였습니다. ")
        
    def saveModel(self, model, filename):
        """
        훈련된 모델을 파일로 저장하는 함수
        
        Args:
            model: 저장할 모델 객체
            filename (str): 저장할 파일명
            
        설명:
        - 훈련된 모델을 나중에 불러와서 예측에 사용할 수 있도록 저장
        - joblib을 사용하여 Python 객체를 효율적으로 저장
        """
        joblib.dump(model, self.save_dir+filename)  # 모델을 파일로 저장 (예: 'gb.model1.v0.0.1.npz')
        
        print("-"*80)
        print("그래디언트 부스팅 모델을 저장하였습니다. ")        
        
    def predict(self, model, x_testset):
        """
        훈련된 모델로 예측을 수행하는 함수
        
        Args:
            model: 예측에 사용할 모델
            x_testset (list): 예측할 입력 데이터
            
        Returns:
            numpy.array: 예측 결과
            
        설명:
        - 훈련된 모델이 새로운 데이터에 대해 예측을 수행
        - 테스트 데이터로 모델의 성능을 평가할 때 사용
        """
        result = model.predict(x_testset)  # 모델로 예측 수행
        
        print("테스트셋으로 예측완료. ")
        
        return result

    def mergeResultset(self, results):
        """
        예측 결과들을 하나의 데이터프레임으로 합치는 함수
        
        Args:
            results (list): [모델1결과, 모델2결과, 모델3결과] 리스트
            
        Returns:
            pandas.DataFrame: 원본 데이터와 예측 결과가 합쳐진 데이터프레임
        """
        df_result = self.make_result_dataframe2(self.xx_test, results[0], results[1], results[2])
        
        return df_result
        
        
    def predictByTestset(self, x_testset):
        """
        3개의 모델로 테스트 데이터에 대해 예측을 수행하는 함수
        
        Args:
            x_testset (list): 예측할 테스트 데이터
            
        Returns:
            pandas.DataFrame: 예측 결과가 포함된 데이터프레임
            
        설명:
        - 3개의 모델을 모두 사용하여 예측 수행
        - 결과를 데이터프레임으로 정리하여 반환
        """
        # 3개의 모델로 각각 예측 수행
        result1 = self.model1.predict(x_testset)  # 업체 투찰률 예측
        result2 = self.model2.predict(x_testset)  # 예가 투찰률 예측
        result3 = self.model3.predict(x_testset)  # 참여 업체 수 예측
        
        print("테스트셋으로 예측완료. ")
        
        # 예측 결과를 데이터프레임으로 정리
        df_result = self.make_result_dataframe2(self.xx_test, result1, result2, result3)
        
        return df_result
    
    
    def saveResultToXls(self, df_result, xls_dir):
        """
        예측 결과를 엑셀 파일로 저장하는 함수
        
        Args:
            df_result (pandas.DataFrame): 저장할 데이터프레임
            xls_dir (str): 저장할 엑셀 파일 경로
            
        설명:
        - 예측 결과를 엑셀 파일로 저장하여 분석할 수 있도록 함
        - openpyxl이 없으면 CSV 파일로 저장
        - 결과1, 결과2 통계를 마지막에 추가
        """
        print("="*80)
        print("예측 결과를 엑셀 파일로 저장 중...")
        print(f"저장 경로: {xls_dir}")
        print(f"데이터 행 수: {len(df_result)}")
        print(f"데이터 열 수: {len(df_result.columns)}")
        
        try:
            # 엑셀 파일로 저장 시도
            df_result.to_excel(
                           xls_dir,  # 저장할 파일 경로
                           sheet_name = 'Sheet1',        # 시트 이름
                           na_rep = 'NaN',               # 결측값 표시
                           float_format = "%.8f",        # 소수점 8자리까지 표시
                           header = True,                # 컬럼명 포함
                           index = True,                 # 인덱스 포함
                           index_label = "id",           # 인덱스 컬럼명
                           startrow = 1,                 # 시작 행
                           startcol = 1,                 # 시작 열
                           freeze_panes = (2, 0)         # 첫 2행 고정 (스크롤 시에도 컬럼명 유지)
                           )
            
            # ===== 결과1, 결과2 통계 추가 =====
            self.addResultStatistics(xls_dir, df_result)
            
            print("="*80)
            print("✅ 예측데이타를 엑셀파일로 저장하였습니다.")
            print(f"📁 저장된 파일: {xls_dir}")
            print(f"📊 데이터 크기: {len(df_result)}행 x {len(df_result.columns)}열")
            print("📈 결과1, 결과2 통계가 추가되었습니다.")
            print("="*80)
        except ModuleNotFoundError:
            # openpyxl이 없으면 CSV 파일로 저장
            csv_dir = xls_dir.replace('.xlsx', '.csv')
            df_result.to_csv(csv_dir, 
                           na_rep = 'NaN', 
                           float_format = "%.8f", 
                           header = True, 
                           index = True, 
                           index_label = "id",
                           encoding='utf-8-sig')
            print("="*80)
            print("⚠️  openpyxl이 없어서 CSV 파일로 저장하였습니다.")
            print(f"📁 저장된 파일: {csv_dir}")
            print(f"📊 데이터 크기: {len(df_result)}행 x {len(df_result.columns)}열")
            print("="*80)
        except Exception as e:
            print(f"❌ 파일 저장 중 오류 발생: {e}")
            # 최후의 수단으로 CSV 파일로 저장
            csv_dir = xls_dir.replace('.xlsx', '.csv')
            df_result.to_csv(csv_dir, 
                           na_rep = 'NaN', 
                           float_format = "%.8f", 
                           header = True, 
                           index = True, 
                           index_label = "id",
                           encoding='utf-8-sig')
            print("="*80)
            print("✅ CSV 파일로 저장하였습니다.")
            print(f"📁 저장된 파일: {csv_dir}")
            print(f"📊 데이터 크기: {len(df_result)}행 x {len(df_result.columns)}열")
            print("="*80)
    
    def addResultStatistics(self, xls_dir, df_result):
        """
        엑셀 파일에 결과1, 결과2 통계를 추가하는 함수
        
        Args:
            xls_dir (str): 엑셀 파일 경로
            df_result (pandas.DataFrame): 결과 데이터프레임
            
        설명:
        - 결과1 컬럼(AD)의 "낙찰" 개수와 비율 계산
        - 결과2 컬럼(AG)의 "낙찰" 개수와 비율 계산
        - 엑셀 파일 마지막에 통계 추가
        """
        try:
            if not OPENPYXL_AVAILABLE:
                raise ImportError("openpyxl이 설치되지 않았습니다.")
            
            # 엑셀 파일 로드
            wb = load_workbook(xls_dir)
            ws = wb.active
            
            # 데이터 행 수 계산 (헤더 제외)
            data_rows = len(df_result)
            start_row = 3  # 데이터 시작 행 (헤더 2행 + 1)
            end_row = start_row + data_rows - 1
            
            # ===== 결과1 통계 추가 =====
            # 결과1 낙찰 개수 (AD 컬럼) - A값여부와 같은 열(AC)에 배치
            stats_row = end_row + 2
            ws[f'AC{stats_row}'] = "=== 결과1 통계 ==="
            ws[f'AC{stats_row}'].font = openpyxl.styles.Font(bold=True, color="0000FF")
            
            # 결과1 낙찰 개수 - 더 안전한 공식 사용
            ws[f'AC{stats_row + 1}'] = "낙찰 개수:"
            count_formula1 = f'=COUNTIF(AD{start_row}:AD{end_row},"낙찰")'
            ws[f'AD{stats_row + 1}'] = count_formula1
            
            # 결과1 낙찰 비율 - 더 안전한 공식 사용
            ws[f'AC{stats_row + 2}'] = "낙찰 비율:"
            rate_formula1 = f'=IF(AD{stats_row + 1}>0,AD{stats_row + 1}/{data_rows}*100,0)'
            ws[f'AD{stats_row + 2}'] = rate_formula1
            
            # ===== 결과2 통계 추가 =====
            # 결과2 낙찰 개수 (AG 컬럼) - 결과1과 동일한 행에 배치 (AF 컬럼에 설명)
            ws[f'AF{stats_row}'] = "=== 결과2 통계 ==="
            ws[f'AF{stats_row}'].font = openpyxl.styles.Font(bold=True, color="00AA00")
            
            # 결과2 낙찰 개수 - 더 안전한 공식 사용
            ws[f'AF{stats_row + 1}'] = "낙찰 개수:"
            count_formula2 = f'=COUNTIF(AG{start_row}:AG{end_row},"낙찰")'
            ws[f'AG{stats_row + 1}'] = count_formula2
            
            # 결과2 낙찰 비율 - 더 안전한 공식 사용
            ws[f'AF{stats_row + 2}'] = "낙찰 비율:"
            rate_formula2 = f'=IF(AG{stats_row + 1}>0,AG{stats_row + 1}/{data_rows}*100,0)'
            ws[f'AG{stats_row + 2}'] = rate_formula2
            
            # ===== 추가 정보 =====
            ws[f'A{stats_row + 4}'] = "=== 요약 정보 ==="
            ws[f'A{stats_row + 4}'].font = openpyxl.styles.Font(bold=True, color="FF0000")
            ws[f'A{stats_row + 5}'] = f"총 데이터 개수: {data_rows}개"
            ws[f'A{stats_row + 6}'] = f"데이터 범위: AD{start_row}:AD{end_row}, AG{start_row}:AG{end_row}"
            
            # 엑셀 파일 저장
            wb.save(xls_dir)
            
            print(f"📊 결과1 통계 추가 완료:")
            print(f"   - 낙찰 개수: AD{stats_row + 1} = {count_formula1}")
            print(f"   - 낙찰 비율: AD{stats_row + 2} = {rate_formula1}")
            print(f"📊 결과2 통계 추가 완료:")
            print(f"   - 낙찰 개수: AG{stats_row + 1} = {count_formula2}")
            print(f"   - 낙찰 비율: AG{stats_row + 2} = {rate_formula2}")
            print(f"📊 통계 위치: {stats_row}행 (데이터 마지막 + 2행)")
            
        except ImportError:
            print("⚠️  openpyxl이 설치되지 않아 통계를 추가할 수 없습니다.")
            print("   pip install openpyxl 명령으로 설치하세요.")
        except Exception as e:
            print(f"❌ 통계 추가 중 오류 발생: {e}")
            print(f"   오류 상세: {str(e)}") 
        
    
    def close(self):
        """
        훈련 과정을 마무리하는 함수
        
        설명:
        - 전체 훈련 과정에 걸린 시간을 계산하여 출력
        - 훈련 완료 메시지 출력
        """
        print("-"*80)
        
        # 전체 실행 시간 계산 및 출력
        print(f"⏱️  전체 실행 시간: {time() - self.t0:.3f}초")
        
        print("="*80)
        print("🎉 머신러닝 훈련 프로세스 완료!")
        if self.excel_file_nm:
            print(f"📁 결과 파일: {self.excel_file_nm}")
            print(f"📂 저장 위치: {self.save_dir}")
        else:
            print("⚠️  엑셀 파일명이 생성되지 않았습니다.")
        print("="*80)

        



def Main():
    """
    머신러닝 모델 훈련의 전체 과정을 실행하는 메인 함수
    
    실행 과정:
    1. 훈련 객체 생성
    2. 데이터 로드 및 전처리
    3. 3개 모델 설정
    4. 각 모델 훈련 및 저장
    5. 예측 수행 및 결과 저장
    6. 엑셀 파일로 결과 출력
    """
    # ===== 1단계: 훈련 객체 생성 =====
    trainer = BidLowerMarginRateTrain()
    
    # ===== 2단계: 데이터 로드 및 전처리 =====
    x_train, x_test, y_train, y_test = trainer.loadTrainsetFromFile('bid_250921_30.csv')  # CSV 파일에서 데이터 로드
    x_trainset, x_testset = trainer.preprocessingXset(x_train, x_test, 'gb_scaler.v2.npz')  # 입력 데이터 정규화
    y_trainset, y_testset = trainer.preprocessingYset(y_train, y_test)  # 출력 데이터 분리
    
    # ===== 3단계: 3개 모델 설정 =====
    models = trainer.setupModels()  # [업체투찰률모델, 예가투찰률모델, 참여업체수모델]
    results = []  # 예측 결과를 저장할 리스트
    
    # ===== 4단계: 각 모델 훈련 및 저장 =====
    for i, model in enumerate(models):
        trainer.trainnng(model, x_trainset, y_trainset[i])  # 모델 훈련
        trainer.saveModel(model, f'gb.model{i+1}.v0.1.1.npz')  # 모델 저장
        result = trainer.predict(model, x_testset)  # 테스트 데이터로 예측
        print(f"모델{i+1} 예측 결과 (처음 50개):")
        print(result[:50])
        print("="*80)
        results.append(result)  # 예측 결과 저장
        
    # ===== 5단계: 결과 정리 및 저장 =====
    print("="*80)
    print("📊 예측 결과를 데이터프레임으로 정리 중...")
    df_result = trainer.mergeResultset(results)  # 예측 결과를 데이터프레임으로 정리
    print(f"✅ 데이터프레임 생성 완료: {len(df_result)}행 x {len(df_result.columns)}열")
    
    print("="*80)
    print("💾 엑셀 파일로 결과 저장 중...")
    trainer.saveResultToXls(df_result, trainer.xlxs_dir)  # 엑셀 파일로 결과 저장
    
    trainer.close()  # 훈련 과정 마무리

    
if __name__ == "__main__":
    """
    스크립트가 직접 실행될 때만 Main() 함수를 호출
    
    설명:
    - 이 파일이 직접 실행될 때만 머신러닝 훈련이 시작됨
    - 다른 파일에서 import할 때는 실행되지 않음
    """
    Main()